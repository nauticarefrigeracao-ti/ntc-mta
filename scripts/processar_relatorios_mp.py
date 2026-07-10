"""
Processador de Relatórios MP – Análise Financeira de Devoluções
===============================================================
Uso:
    python scripts/processar_relatorios_mp.py
    python scripts/processar_relatorios_mp.py --pasta tmp_csvs/ --output reports/
    python scripts/processar_relatorios_mp.py --pasta tmp_csvs/ --force

O script:
  1. Varre a pasta em busca de arquivos after_collection (Excel/CSV)
  2. Ingere os novos relatórios no Neon (tabela mp_transactions)
  3. Cruza com orders, ml_devolucoes, order_items, tiny_sku_costs
  4. Gera relatório XLSX multi-aba (PT-BR, formatado)
  5. Gera dashboard HTML interativo com gráficos Plotly
  6. Imprime resumo no console
"""
from __future__ import annotations

import argparse
import json
import sys
import warnings
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np

warnings.filterwarnings("ignore")
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ─── paths ───────────────────────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from src.db.connection import get_db_connection
from src.services.mp_ingestion import scan_folder, list_imported, BUCKET_MAP
from src.api import ml_client

# ─── constantes ──────────────────────────────────────────────────────────────
TODAY        = datetime.now().strftime("%Y-%m-%d")
DEFAULT_IN   = ROOT / "tmp_csvs"
DEFAULT_OUT  = ROOT / "reports"
BRL          = lambda v: f"R$ {float(v or 0):,.2f}"
PCT          = lambda v: f"{float(v or 0):.1f}%"
SEP          = "=" * 70

BUCKET_CORES: dict[str, str] = {
    "Protegido ML":     "#27AE60",
    "Mediação ML":      "#2980B9",
    "Perda Confirmada": "#E74C3C",
    "Reembolso Direto": "#F39C12",
    "Administrativo":   "#95A5A6",
    "Outro":            "#BDC3C7",
}

# Labels em português para exibição em relatórios e gráficos (linguagem de negócio)
STATUS_DETAIL_LABEL: dict[str, str] = {
    "bpp_refunded":           "Reembolso Automático (Proteção ML)",
    "bpp_covered":            "Cobertura Automática (Proteção ML)",
    "partially_bpp_refunded": "Reembolso Parcial (Proteção ML)",
    "partially_bpp_covered":  "Cobertura Parcial (Proteção ML)",
    "reconciled":             "Mediação Encerrada — ML Arcou",
    "compensated":            "Indenização pelo ML",
    "not_reconciled":         "Não Conciliado — Perda do Vendedor",
    "refunded":               "Reembolso pelo Vendedor",
    "by_admin":               "Decisão Administrativa ML",
    "ppv_covered_melienvio":  "Cobertura Envios ML (PPV)",
}
STATUS_DETAIL_CORES: dict[str, str] = {
    "bpp_refunded":           "#27AE60",
    "bpp_covered":            "#2ECC71",
    "partially_bpp_refunded": "#7DCEA0",
    "partially_bpp_covered":  "#A9DFBF",
    "reconciled":             "#2980B9",
    "compensated":            "#5DADE2",
    "not_reconciled":         "#E74C3C",
    "refunded":               "#F39C12",
    "by_admin":               "#95A5A6",
    "ppv_covered_melienvio":  "#48C9B0",
}

# ─── SQL de análise ───────────────────────────────────────────────────────────
_SQL_ANALISE = """
SELECT
    t.order_id                          AS mp_order_id,
    t.data_criacao                      AS mp_data_criacao,
    t.status_detail,
    t.categoria,
    t.valor                             AS mp_valor,
    t.reason_id                         AS mp_motivo,
    t.source_file,
    -- orders
    o.data_venda,
    o.total_brl                         AS csv_total,
    o.receita_produtos_brl              AS csv_receita_produtos,
    o.receita_envio_brl                 AS csv_receita_envio,
    o.tarifa_venda_impostos_brl         AS taxa_ml_csv,
    o.tarifas_envio_brl                 AS tarifa_frete_csv,
    o.cancelamentos_reembolsos_brl      AS cancelamentos_csv,
    o.dinheiro_liberado                 AS dinheiro_liberado_csv,
    -- ml claims
    m.claim_id,
    m.claim_status,
    m.claim_type,
    m.reason_label                      AS motivo_ml,
    m.order_total                       AS perda_bruta,
    m.sale_fee                          AS taxa_ml_retida,
    m.amount_refunded                   AS recuperado_ml,
    -- items
    oi.sku,
    oi.preco_unitario                   AS preco_unitario,
    oi.unidades,
    -- tiny CMV
    tc.preco_custo                      AS cmv_unitario,
    tc.nome                             AS produto_tiny
FROM mp_transactions t
LEFT JOIN orders          o  ON o.order_id       = t.order_id::text
LEFT JOIN ml_devolucoes   m  ON m.order_id        = t.order_id
                             AND m.claim_type      = 'mediations'
                             AND m.claim_status    = 'closed'
LEFT JOIN order_items     oi ON oi.order_id        = t.order_id::text
LEFT JOIN tiny_sku_costs  tc ON upper(tc.sku)      = upper(oi.sku)
ORDER BY t.data_criacao DESC
"""

_SQL_PORTFOLIO = """
SELECT
    o.order_id,
    o.data_venda,
    o.estado,
    o.descricao_status,
    o.total_brl,
    o.receita_produtos_brl,
    o.receita_envio_brl,
    o.tarifa_venda_impostos_brl,
    o.tarifas_envio_brl,
    o.cancelamentos_reembolsos_brl,
    o.dinheiro_liberado,
    m.order_total        AS perda_bruta,
    m.sale_fee           AS taxa_ml_retida,
    m.amount_refunded    AS recuperado_ml,
    m.reason_label       AS motivo_ml,
    oi.sku,
    tc.preco_custo       AS cmv
FROM orders o
LEFT JOIN ml_devolucoes m  ON m.order_id = o.order_id::BIGINT
                           AND m.claim_type = 'mediations'
                           AND m.claim_status = 'closed'
LEFT JOIN order_items oi   ON oi.order_id = o.order_id
LEFT JOIN tiny_sku_costs tc ON upper(tc.sku) = upper(oi.sku)
"""

# ─── parse & normalização ─────────────────────────────────────────────────────

def _num(df: pd.DataFrame, col: str) -> pd.Series:
    return pd.to_numeric(df[col], errors="coerce").fillna(0.0) if col in df.columns else pd.Series(0.0, index=df.index)


def _load_and_enrich(conn) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Retorna (df_mp, df_portfolio)."""
    df_mp   = pd.read_sql(_SQL_ANALISE,    conn)
    df_port = pd.read_sql(_SQL_PORTFOLIO,  conn)

    # Normalizar numéricos
    for c in ["mp_valor","csv_total","csv_receita_produtos","csv_receita_envio",
              "taxa_ml_csv","tarifa_frete_csv","cancelamentos_csv","dinheiro_liberado_csv",
              "perda_bruta","taxa_ml_retida","recuperado_ml","preco_unitario","cmv_unitario"]:
        df_mp[c] = _num(df_mp, c)

    for c in ["total_brl","receita_produtos_brl","receita_envio_brl",
              "tarifa_venda_impostos_brl","tarifas_envio_brl",
              "cancelamentos_reembolsos_brl","dinheiro_liberado",
              "perda_bruta","taxa_ml_retida","recuperado_ml","cmv"]:
        df_port[c] = _num(df_port, c)

    df_mp["data_criacao"] = pd.to_datetime(df_mp["mp_data_criacao"], errors="coerce", utc=True)
    df_mp["mes"]          = df_mp["data_criacao"].dt.to_period("M").astype(str)

    df_port["data_venda"] = pd.to_datetime(df_port["data_venda"], errors="coerce", utc=True)
    df_port["mes"]        = df_port["data_venda"].dt.to_period("M").astype(str)
    df_port["semestre"]   = df_port["data_venda"].apply(
        lambda d: f"{d.year}-S1" if pd.notna(d) and d.month <= 6 else (f"{d.year}-S2" if pd.notna(d) else pd.NA)
    )

    # Fórmulas financeiras do portfólio
    df_port["receita_bruta"]  = df_port["receita_produtos_brl"] + df_port["receita_envio_brl"].clip(lower=0)
    df_port["taxa_ml"]        = df_port["tarifa_venda_impostos_brl"].clip(upper=0).abs()
    df_port["tarifa_frete"]   = df_port["tarifas_envio_brl"].clip(upper=0).abs()
    df_port["cancelamentos"]  = df_port["cancelamentos_reembolsos_brl"].clip(upper=0).abs()
    df_port["perda_liquida"]  = (df_port["perda_bruta"] - df_port["recuperado_ml"]).clip(lower=0)
    df_port["impacto_cmv"]    = df_port.apply(
        lambda r: r["cmv"] + r["taxa_ml_retida"] - r["recuperado_ml"]
                  if r["cmv"] > 0 and r["perda_bruta"] > 0 else None,
        axis=1,
    )

    # Fórmulas para df_mp
    df_mp["perda_liquida"] = (df_mp["perda_bruta"] - df_mp["recuperado_ml"]).clip(lower=0)
    df_mp["pct_recuperado"] = (
        df_mp["recuperado_ml"] / df_mp["perda_bruta"].where(df_mp["perda_bruta"] > 0) * 100
    ).round(1)

    # ── Conciliação de estado — ML API (vivo) + relatório collection (offline) ──
    # Regras validadas contra a plataforma Meli (amostras por pedido):
    #  1. Pagamento 'accredited' = dinheiro CREDITADO ao vendedor e nenhum
    #     reembolso ao comprador → a disputa encerrou sem perda ("mantido").
    #     Ex.: comprador cancela a reclamação → "Te demos o dinheiro desta venda".
    #  2. Taxa ML só é perdida quando a linha terminou em perda real — nas
    #     revertidas/canceladas o Meli estorna a tarifa ("Cancelamento de tarifa").
    state: dict[str, str] = {}
    try:
        df_coll = pd.read_sql(
            "SELECT order_id::text AS oid, MAX(status_detail) AS sd, "
            "SUM(COALESCE(valor_devolvido,0)) AS vd "
            "FROM mp_collection WHERE order_id IS NOT NULL GROUP BY 1", conn)
        for _, r in df_coll.iterrows():
            # só considera 'accredited' se o collection confirma zero devolvido
            if str(r["sd"]) == "accredited" and float(r["vd"] or 0) <= 0.005:
                state[str(r["oid"])] = "accredited"
    except Exception:
        pass
    try:
        df_api = pd.read_sql(
            "SELECT order_id::text AS oid, api_pay_detail FROM mp_validation_results "
            "WHERE api_pay_detail IS NOT NULL AND api_pay_detail <> ''", conn)
        for _, r in df_api.iterrows():
            state[str(r["oid"])] = str(r["api_pay_detail"])  # API sobrescreve (mais fresca)
    except Exception:
        pass

    acc_ids = {oid for oid, s in state.items() if s == "accredited"}

    for df, idcol in ((df_port, "order_id"), (df_mp, "mp_order_id")):
        oid_str = df[idcol].astype(str)
        manteve = oid_str.isin(acc_ids) & (df["perda_bruta"] > 0)
        df["mantido"] = 0.0
        df.loc[manteve, "mantido"] = (
            df.loc[manteve, "perda_bruta"] - df.loc[manteve, "recuperado_ml"]
        ).clip(lower=0)
        df.loc[manteve, "perda_liquida"] = 0.0
        df["taxa_ml_retida_raw"] = df["taxa_ml_retida"]
        df["taxa_ml_retida"] = df["taxa_ml_retida"].where(df["perda_liquida"] > 0.005, 0.0)

    return df_mp, df_port


# ─── KPIs ─────────────────────────────────────────────────────────────────────

def _kpis(df_mp: pd.DataFrame, df_port: pd.DataFrame) -> dict:
    # Portfólio completo
    total_pedidos      = df_port["order_id"].nunique()
    receita_bruta_tot  = df_port["receita_bruta"].sum()
    taxa_ml_tot        = df_port["taxa_ml"].sum()
    tarifa_frete_tot   = df_port["tarifa_frete"].sum()
    frete_cobrado_tot  = df_port["receita_envio_brl"].clip(lower=0).sum()
    cancelamentos_tot  = df_port["cancelamentos"].sum()
    receita_liq_tot    = df_port["total_brl"].sum()

    # Devoluções (só pedidos com claim)
    df_dev             = df_port[df_port["perda_bruta"] > 0].drop_duplicates("order_id")
    perda_bruta_tot    = df_dev["perda_bruta"].sum()
    recuperado_tot     = df_dev["recuperado_ml"].sum()
    taxa_retida_tot    = df_dev["taxa_ml_retida"].sum()
    perda_liq_tot      = df_dev["perda_liquida"].sum()
    mantido_tot        = df_dev["mantido"].sum() if "mantido" in df_dev.columns else 0.0
    pct_recup          = recuperado_tot / perda_bruta_tot * 100 if perda_bruta_tot else 0

    # Relatório MP (transações ingeridas)
    mp_total_tx        = len(df_mp)
    mp_pedidos_unicos  = df_mp["mp_order_id"].nunique()
    mp_valor_total     = df_mp["mp_valor"].sum()

    return {
        "total_pedidos":       total_pedidos,
        "receita_bruta":       receita_bruta_tot,
        "taxa_ml":             taxa_ml_tot,
        "tarifa_frete":        tarifa_frete_tot,
        "frete_cobrado":       frete_cobrado_tot,
        "saldo_frete":         frete_cobrado_tot - tarifa_frete_tot,
        "cancelamentos":       cancelamentos_tot,
        "receita_liquida":     receita_liq_tot,
        "dev_pedidos":         len(df_dev),
        "dev_pct_portfolio":   len(df_dev) / total_pedidos * 100 if total_pedidos else 0,
        "perda_bruta":         perda_bruta_tot,
        "recuperado_ml":       recuperado_tot,
        "taxa_retida":         taxa_retida_tot,
        "perda_liquida":       perda_liq_tot,
        "mantido_vendedor":    mantido_tot,
        "pct_recuperado":      pct_recup,
        "mp_total_tx":         mp_total_tx,
        "mp_pedidos_unicos":   mp_pedidos_unicos,
        "mp_valor_total":      mp_valor_total,
    }


# ─── Gráficos Plotly (HTML interativo) ────────────────────────────────────────

def _build_html_dashboard(
    df_mp: pd.DataFrame,
    df_port: pd.DataFrame,
    kpi: dict,
    output_path: Path,
    validation: dict | None = None,
) -> None:
    """Dashboard interativo — 3 categorias definidas, filtro global JS, cards clicáveis."""
    import json as _json
    try:
        import plotly.graph_objects as go
    except ImportError:
        print("  ⚠ plotly não instalado – dashboard HTML ignorado")
        return

    # ═══════════════════════════════════════════════════════════════════════════
    # DEFINIÇÕES DE NEGÓCIO
    # ───────────────────────────────────────────────────────────────────────────
    # CANCELAMENTO  = pedido cancelado antes/sem entrega física
    #   → MP: status_detail = 'refunded' (reembolso direto pelo vendedor)
    #   → Impacto: geralmente zero; pode perder taxa ML se já processado
    #
    # DEVOLUÇÃO     = produto entregue e devolvido pelo comprador
    #   → ML cobre automaticamente via BPP (Buyer Protection Program)
    #   → MP: bpp_refunded, bpp_covered, partially_bpp_refunded, partially_bpp_covered
    #   → Impacto: frete reverso cobrado do vendedor + possível perda do produto
    #
    # RECLAMAÇÃO    = disputa formal aberta pelo comprador via ML
    #   → ML arbitra com base em fotos/rastreio/laudos
    #   → MP: reconciled, compensated, not_reconciled, by_admin
    #   → ml_devolucoes: claim_type = 'mediations'
    #   → Impacto: variável — ML decide quem paga
    # ═══════════════════════════════════════════════════════════════════════════

    SD_DEVOLUCAO    = {"bpp_refunded","bpp_covered","partially_bpp_refunded","partially_bpp_covered"}
    SD_RECLAMACAO   = {"reconciled","compensated","not_reconciled","by_admin"}
    SD_CANCELAMENTO = {"refunded"}

    # ── Base de dados ─────────────────────────────────────────────────────────
    df_port_u = df_port.drop_duplicates("order_id").copy()
    df_port_u["data_venda"] = pd.to_datetime(df_port_u["data_venda"], utc=True, errors="coerce")
    df_port_u["dia"] = df_port_u["data_venda"].dt.strftime("%Y-%m-%d").fillna("")

    df_dev = df_port_u[df_port_u["perda_bruta"] > 0].copy()
    df_dev["dia"] = df_dev["data_venda"].dt.strftime("%Y-%m-%d").fillna("")

    df_mp_c = df_mp.copy()
    df_mp_c["data_criacao"] = pd.to_datetime(df_mp_c["data_criacao"], errors="coerce", utc=True)
    df_mp_c["dia"] = df_mp_c["data_criacao"].dt.strftime("%Y-%m-%d").fillna("")
    df_mp_c["mp_valor_f"] = pd.to_numeric(df_mp_c.get("mp_valor", 0), errors="coerce").fillna(0.0)

    df_cancel   = df_mp_c[df_mp_c["status_detail"].isin(SD_CANCELAMENTO)].copy()
    df_devol    = df_mp_c[df_mp_c["status_detail"].isin(SD_DEVOLUCAO)].copy()
    df_recl     = df_mp_c[df_mp_c["status_detail"].isin(SD_RECLAMACAO)].copy()
    df_mediacao = df_recl[df_recl["status_detail"].isin({"reconciled","compensated"})].copy()
    df_nc       = df_recl[df_recl["status_detail"] == "not_reconciled"].copy()

    # ── CANCELAMENTOS REAIS (fonte: orders.estado, não o arquivo MP) ──────────
    # O relatório MP after_collection não traz linhas 'refunded' — o card ficava
    # zerado. A fonte correta é o estado do pedido na base ML (pedido cancelado
    # antes da entrega). O reembolso ao comprador está em cancelamentos_reembolsos_brl.
    _CANCEL_RE = r"(?i)^(cancelada|pacote cancelado|venda cancelada|você cancelou)"
    _estado = df_port_u.get("estado", pd.Series("", index=df_port_u.index)).astype(str)
    df_cancel_real = df_port_u[_estado.str.match(_CANCEL_RE, na=False)].copy()
    df_cancel_real["valor_reemb"] = df_cancel_real["cancelamentos_reembolsos_brl"].clip(upper=0).abs()
    df_cancel_real["resid"]       = df_cancel_real["total_brl"]
    df_cancel_real["anom"]        = (df_cancel_real["resid"].abs() > 0.10).astype(int)

    def _tipo_cancel(estado: str, motivo: str) -> str:
        e = (estado or "").lower()
        m = (motivo or "").lower()
        if (e.startswith("cancelada pelo comprador")
                or "comprou cancelou" in m or "comprador se arrependeu" in m
                or "comprador não concorda" in m or m.startswith("cancelou")):
            return "Comprador cancelou"
        if (e.startswith("pacote cancelado") or "cancelada pelo mercado" in e
                or m.startswith("tivemos que cancelar")):
            return "Mercado Livre cancelou"
        if e.startswith("você cancelou"):
            return "Vendedor cancelou"
        return "Cancelado (ver motivo)"

    df_cancel_real["tipo"] = [
        _tipo_cancel(e, m)
        for e, m in zip(df_cancel_real["estado"].astype(str), df_cancel_real["descricao_status"].astype(str))
    ]
    df_cancel_real["motivo"] = df_cancel_real["descricao_status"].astype(str).str.slice(0, 70)

    ids_prej = (
        set(df_dev["order_id"].astype(str)) |
        set(df_devol["mp_order_id"].astype(str)) |
        set(df_recl["mp_order_id"].astype(str))
    )
    frete_prej  = df_port_u[df_port_u["order_id"].astype(str).isin(ids_prej)]["tarifa_frete"].sum()
    df_revert   = df_dev[(df_dev["perda_bruta"] > 0) & (df_dev["recuperado_ml"] >= df_dev["perda_bruta"] * 0.90)].copy()

    # ── Dados diários para o filtro JS ────────────────────────────────────────
    def _daily(df, gcol, aggs):
        if df.empty or gcol not in df.columns:
            return pd.DataFrame(columns=[gcol] + list(aggs.keys()))
        return df.groupby(gcol).agg(**aggs).reset_index().sort_values(gcol)

    daily_port   = _daily(df_port_u,"dia",{"pedidos":pd.NamedAgg("order_id","count"),"frete":pd.NamedAgg("tarifa_frete","sum"),"receita":pd.NamedAgg("total_brl","sum")})
    daily_dev    = _daily(df_dev,"dia",{"count":pd.NamedAgg("order_id","count"),"perda":pd.NamedAgg("perda_bruta","sum"),"recuperado":pd.NamedAgg("recuperado_ml","sum"),"taxa":pd.NamedAgg("taxa_ml_retida","sum"),"liq":pd.NamedAgg("perda_liquida","sum"),"mantido":pd.NamedAgg("mantido","sum")})
    daily_devol  = _daily(df_devol,"dia",{"count":pd.NamedAgg("mp_order_id","count"),"valor":pd.NamedAgg("mp_valor_f","sum")})
    daily_recl   = _daily(df_recl,"dia",{"count":pd.NamedAgg("mp_order_id","count"),"valor":pd.NamedAgg("mp_valor_f","sum")})
    daily_cancel = _daily(df_cancel,"dia",{"count":pd.NamedAgg("mp_order_id","count"),"valor":pd.NamedAgg("mp_valor_f","sum")})
    daily_cancel_real = _daily(df_cancel_real,"dia",{"count":pd.NamedAgg("order_id","count"),"valor":pd.NamedAgg("valor_reemb","sum"),"anom":pd.NamedAgg("anom","sum")})

    def _rows(df, cols, limit=500):
        out = []
        for _, r in df[cols].head(limit).iterrows():
            d = {}
            for c in cols:
                v = r[c]
                if isinstance(v, pd.Timestamp): d[c] = str(v)[:10]
                elif isinstance(v, float) and v != v: d[c] = None
                elif isinstance(v, float): d[c] = round(v, 2)
                elif hasattr(v, "item"): d[c] = v.item()
                else: d[c] = v
            out.append(d)
        return out

    def _df2js(df):
        return _rows(df, list(df.columns), limit=10000)

    rec_c  = ["order_id","dia","sku","motivo_ml","perda_bruta","taxa_ml_retida","recuperado_ml","perda_liquida"]
    mp_c   = ["mp_order_id","dia","status_detail","mp_valor_f"]
    rev_c  = ["order_id","dia","sku","perda_bruta","recuperado_ml"]
    can_c  = ["order_id","dia","tipo","motivo","valor_reemb","resid"]

    recl_modal     = _rows(df_dev,      rec_c)
    devol_modal    = _rows(df_devol,    mp_c)
    cancel_modal   = _rows(df_cancel_real.sort_values("data_venda", ascending=False), can_c, limit=20000)
    mediacao_modal = _rows(df_mediacao, mp_c)
    nc_modal       = _rows(df_nc,       mp_c)
    revert_modal   = _rows(df_revert,   rev_c)

    # ── Estado atual API ML (cache mp_validation_results) ─────────────────────
    try:
        _vc = get_db_connection()
        df_val = pd.read_sql(
            "SELECT order_id::text AS oid, api_pay_detail, api_total, concorda_ml, "
            "to_char(validated_at AT TIME ZONE 'America/Sao_Paulo','DD/MM/YY HH24:MI') AS val_em "
            "FROM mp_validation_results", _vc
        )
        val_map = {str(r["oid"]): r.to_dict() for _, r in df_val.iterrows()}
        _vc.close()
    except Exception:
        val_map = {}

    # ── Classificação de negócio + tradução de status ──────────────────────────
    # RECLAMAÇÕES: vermelho=perda, verde=revertida/mantida, amarelo=parcial
    for r in recl_modal:
        liq = float(r.get("perda_liquida") or 0)
        vr = val_map.get(str(r.get("order_id","")))
        r["estado_api"] = str(vr.get("api_pay_detail") or "—") if vr else "—"
        r["api_em"]     = str(vr.get("val_em") or "—") if vr else "—"
        if r["estado_api"] == "accredited" and liq <= 1:
            # conciliado com o Meli: pagamento creditado, sem reembolso ao comprador
            r["situacao"], r["classe"] = "🟢 Sem Perda — venda creditada", "ok"
        else:
            r["situacao"] = "🔴 Perda" if liq > 5 else ("🟢 Revertida" if liq <= 1 else "🟡 Parcial")
            r["classe"]   = "bad" if liq > 5 else ("ok" if liq <= 1 else "warn")

    # DEVOLUÇÕES + MEDIAÇÕES: traduz status, enriquece com API
    for r in devol_modal + mediacao_modal:
        sd = str(r.get("status_detail",""))
        r["status_detail"] = STATUS_DETAIL_LABEL.get(sd, sd)
        vr = val_map.get(str(r.get("mp_order_id","")))
        r["estado_api"] = str(vr.get("api_pay_detail") or "—") if vr else "—"
        r["api_em"]     = str(vr.get("val_em") or "—") if vr else "—"
        valor = float(r.get("mp_valor_f") or 0)
        r["classe"]   = "ok" if valor > 0.01 else "warn"
        r["situacao"] = "✔ ML Arcou" if valor > 0.01 else "⚠ Verificar"

    # NÃO CONCILIADOS: sempre perda
    for r in nc_modal:
        sd = str(r.get("status_detail",""))
        r["status_detail"] = STATUS_DETAIL_LABEL.get(sd, sd)
        r["classe"]   = "bad"
        r["situacao"] = "🔴 Perda Confirmada"
        vr = val_map.get(str(r.get("mp_order_id","")))
        r["estado_api"] = str(vr.get("api_pay_detail") or "—") if vr else "—"
        r["api_em"]     = str(vr.get("val_em") or "—") if vr else "—"

    # CANCELAMENTOS: o pedido cancelado DEVE zerar — saldo residual ≠ 0 é anomalia
    for r in cancel_modal:
        resid = float(r.get("resid") or 0)
        reemb = float(r.get("valor_reemb") or 0)
        if abs(resid) > 0.10:
            r["classe"], r["situacao"] = "warn", "⚠ ANOMALIA — saldo residual não zerou"
        elif reemb > 0.005:
            r["classe"], r["situacao"] = "", "✔ Reembolsado — pedido zerado"
        else:
            r["classe"], r["situacao"] = "", "✔ Cancelado sem movimentação"
        vr = val_map.get(str(r.get("order_id","")))
        r["estado_api"] = str(vr.get("api_pay_detail") or "—") if vr else "—"
        r["api_em"]     = str(vr.get("val_em") or "—") if vr else "—"

    # REVERTIDOS: melhor caso
    for r in revert_modal:
        r["classe"]   = "ok"
        r["situacao"] = "🟢 Revertida — ML Cobriu"
        vr = val_map.get(str(r.get("order_id","")))
        r["estado_api"] = str(vr.get("api_pay_detail") or "—") if vr else "—"
        r["api_em"]     = str(vr.get("val_em") or "—") if vr else "—"

    # Frete diário apenas dos pedidos com prejuízo (para card dinâmico)
    df_prej_u = df_port_u[df_port_u["order_id"].astype(str).isin(ids_prej)].copy()
    daily_frete_p = _daily(df_prej_u, "dia", {"frete_prej": pd.NamedAgg("tarifa_frete","sum")})

    embedded = _json.dumps({
        "daily_port":    _df2js(daily_port),
        "daily_dev":     _df2js(daily_dev),
        "daily_devol":   _df2js(daily_devol),
        "daily_recl":    _df2js(daily_recl),
        "daily_cancel":  _df2js(daily_cancel),
        "daily_cancel_real": _df2js(daily_cancel_real),
        "daily_frete_p": _df2js(daily_frete_p),
        "recl_modal":    recl_modal,
        "devol_modal":   devol_modal,
        "cancel_modal":  cancel_modal,
        "mediacao_modal":mediacao_modal,
        "nc_modal":      nc_modal,
        "revert_modal":  revert_modal,
        "d_start": str(df_port_u["dia"].min()),
        "d_end":   str(df_port_u["dia"].max()),
    }, ensure_ascii=False, default=str)

    # ── Gráfico de barras ─────────────────────────────────────────────────────
    fig_bar = go.Figure()
    dev_mes = (
        df_dev.assign(mes=df_dev["data_venda"].dt.strftime("%Y-%m"))
        .groupby("mes").agg(rec=("recuperado_ml","sum"),liq=("perda_liquida","sum"),taxa=("taxa_ml_retida","sum"))
        .reset_index().sort_values("mes")
    )
    if not dev_mes.empty:
        for col,cor,lbl in [("rec","#27AE60","Recuperado pelo ML"),("liq","#E74C3C","Perda Real do Vendedor"),("taxa","#F39C12","Taxa ML Não Devolvida")]:
            fig_bar.add_trace(go.Bar(x=dev_mes["mes"].tolist(),y=dev_mes[col].tolist(),name=lbl,marker_color=cor,hovertemplate="R$ %{y:,.2f}<extra>"+lbl+"</extra>"))
    fig_bar.update_layout(
        barmode="stack", height=330,
        xaxis=dict(type="date", showgrid=True, gridcolor="#EEE",
                   rangeslider=dict(visible=True, thickness=0.05),
                   rangeselector=dict(bgcolor="#E8F4FD", activecolor="#1F4E79",
                                      buttons=[dict(count=3,label="3m",step="month",stepmode="backward"),
                                               dict(count=6,label="6m",step="month",stepmode="backward"),
                                               dict(step="all",label="Tudo")])),
        yaxis=dict(title="R$",showgrid=True,gridcolor="#EEE"),
        plot_bgcolor="white",paper_bgcolor="white",
        legend=dict(orientation="h",y=-0.35),
        font=dict(family="Segoe UI, Arial",size=11),
        margin=dict(t=15,b=80),
    )

    # ── Gráfico donut ─────────────────────────────────────────────────────────
    fig_pie = go.Figure()
    mp_pie = df_mp_c.dropna(subset=["status_detail"])
    mp_pie = mp_pie[mp_pie["status_detail"].astype(str).str.lower().ne("nan")]
    if not mp_pie.empty:
        agg_pie = mp_pie.groupby("status_detail")["mp_valor_f"].sum().reset_index().sort_values("mp_valor_f",ascending=False)
        fig_pie.add_trace(go.Pie(
            labels=[STATUS_DETAIL_LABEL.get(s,s) for s in agg_pie["status_detail"].tolist()],
            values=agg_pie["mp_valor_f"].tolist(), hole=0.52,
            marker_colors=[STATUS_DETAIL_CORES.get(s,"#BDC3C7") for s in agg_pie["status_detail"].tolist()],
            textinfo="percent", sort=True,
            hovertemplate="%{label}<br>R$ %{value:,.2f}<br>%{percent}<extra></extra>",
        ))
    fig_pie.update_layout(height=330,paper_bgcolor="white",
                          font=dict(family="Segoe UI, Arial",size=10),
                          legend=dict(font=dict(size=9),orientation="v",x=1.0),
                          margin=dict(t=5,b=5,l=5,r=5))

    c_bar = fig_bar.to_html(full_html=False, include_plotlyjs="cdn", config={"displayModeBar": True})
    c_pie = fig_pie.to_html(full_html=False, include_plotlyjs=False, config={"displayModeBar": False})

    # ── Valores estáticos iniciais ─────────────────────────────────────────────
    pct_color = "#27AE60" if kpi["pct_recuperado"] >= 70 else "#E67E22" if kpi["pct_recuperado"] >= 40 else "#E74C3C"
    d_start   = str(df_port_u["dia"].min())
    d_end     = str(df_port_u["dia"].max())

    val_badge = ""
    if validation and validation.get("verificados", 0) > 0:
        v = validation
        pv = v["pct_paridade"]
        bc = "#27AE60" if pv >= 90 else "#E67E22" if pv >= 70 else "#E74C3C"
        val_badge = f'<div class="val-badge" style="border-color:{bc};color:{bc}">✓ Conciliação ML API: {pv:.1f}% ({v["concordantes"]:,}/{v["verificados"]:,} | {v.get("com_cmv",0):,} com CMV Tiny)</div>'

    # ── Modais ────────────────────────────────────────────────────────────────
    def _modal(mid, title, note, heads):
        ths = "".join(f"<th>{h}</th>" for h in heads)
        return f"""<dialog id="dlg-{mid}"><div class="dlg-head"><h3>{title}&nbsp;<span style="font-size:11px;opacity:.7">(<span id="cnt-{mid}">0</span>)</span></h3><button class="close-btn" onclick="this.closest('dialog').close()">✕</button></div><div class="dlg-note">{note}</div><div class="dlg-body"><table><thead><tr>{ths}</tr></thead><tbody id="tbody-{mid}"></tbody></table></div></dialog>"""

    RH = ["Nº Pedido","Data","Produto (SKU)","Motivo da Reclamação","Val. Original (R$)","Recuperado ML (R$)","Taxa Retida (R$)","Perda Real (R$)","Situação","Estado Atual (ML API)","Validado em"]
    MH = ["Nº Pedido","Data","Tipo de Resolução","Valor (R$)","Situação","Estado Atual (ML API)","Validado em"]
    CH = ["Nº Pedido","Data","Quem Cancelou","Motivo","Reembolsado (R$)","Saldo Residual (R$)","Situação ⚠","Estado Atual (ML API)","Validado em"]
    VH = ["Nº Pedido","Data","Produto (SKU)","Val. Original (R$)","Recuperado (R$)","Situação","Estado Atual (ML API)"]
    modals = (
        _modal("reclamacoes","Reclamações — Mediações ML","🔴 Vermelho = perda real &nbsp;|  🟡 Amarelo = parcial &nbsp;|  🟢 Verde = revertida &nbsp;|  'Estado Atual' = status vivo na API ML",RH) +
        _modal("devolucoes","Devoluções — Proteção Automática ML (BPP)","ML debitou automaticamente. Confira coluna 'Estado Atual' para status ao vivo na API.",MH) +
        _modal("cancelamentos","Cancelamentos — Pedidos Cancelados (fonte: base de pedidos ML)","Pedido cancelado antes da entrega. O reembolso ao comprador DEVE zerar o pedido — linha amarela = saldo residual diferente de 0 (anomalia, verificar).",CH) +
        _modal("mediacao","Mediações Conciliadas — ML Cobriu","ML arbitrou e cobriu o valor. reconciled = encerrado | compensated = indenizado.",MH) +
        _modal("nao_conciliado","Não Conciliados — Perda Confirmada","🔴 Estas são as reclamações que o vendedor perdeu. Valor saiu da conta sem recuperação.",MH) +
        _modal("revertidos","Revertidos do Negativo — ML Cobriu Tudo","🟢 Melhor cenário: devolviram o produto, ML pagou a venda e o vendedor ficou no zero ou positivo.",VH)
    )

    # ── HTML ──────────────────────────────────────────────────────────────────
    brl = lambda v: f"R$ {float(v or 0):,.2f}".replace(",","X").replace(".",",").replace("X",".")
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    med_c = len(df_mediacao); med_v = df_mediacao["mp_valor_f"].sum()
    nc_c  = len(df_nc);       nc_v  = df_nc["mp_valor_f"].sum()

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="utf-8"><title>Painel de Devoluções</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',Arial,sans-serif;background:#EEF2F7;color:#222}}
.hdr{{background:#1F4E79;color:#fff;padding:15px 22px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:10px}}
.hdr h1{{font-size:18px;font-weight:700}}.hdr .sub{{font-size:11px;opacity:.75;margin-top:2px}}
.val-badge{{border:1.5px solid;border-radius:20px;padding:4px 14px;font-size:11px;font-weight:600;background:#fff;flex-shrink:0}}
.fbar{{background:#1a3a5c;padding:9px 22px;display:flex;align-items:center;gap:8px;flex-wrap:wrap}}
.fbar label{{color:#90B8D8;font-size:11px;font-weight:600;letter-spacing:.5px}}
.btn-r{{background:rgba(255,255,255,.1);border:none;color:#fff;padding:4px 11px;border-radius:6px;font-size:12px;cursor:pointer;transition:background .15s}}
.btn-r:hover{{background:#2980B9}}.btn-r.active{{background:#2980B9;font-weight:700}}
.fbar input[type=date]{{background:rgba(255,255,255,.12);border:1px solid rgba(255,255,255,.25);color:#fff;padding:3px 8px;border-radius:6px;font-size:11px}}
.fbar input[type=date]::-webkit-calendar-picker-indicator{{filter:invert(1)}}
.fperiod{{color:#7EC8E3;font-size:10px;margin-left:6px}}
.sec-hdr{{padding:12px 22px 6px;display:flex;align-items:center;gap:10px;margin-top:14px}}
.sec-hdr h2{{font-size:13px;font-weight:700;color:#1F4E79;white-space:nowrap}}
.sec-hdr hr{{flex:1;border:none;border-top:1.5px solid #C5D8ED}}
.sec-note{{background:#E8EFF8;padding:6px 22px;font-size:10.5px;color:#555;margin-bottom:2px;line-height:1.5}}
.sec-note b{{color:#1F4E79}}
.cards{{display:flex;flex-wrap:wrap;gap:12px;padding:8px 22px 14px}}
.card{{background:#fff;border-radius:10px;padding:14px 16px;min-width:155px;flex:1;max-width:215px;
       box-shadow:0 1px 3px rgba(0,0,0,.08);cursor:pointer;transition:all .15s;
       position:relative;border:2px solid #EEE}}
.card:hover{{box-shadow:0 5px 18px rgba(0,0,0,.14);transform:translateY(-2px);border-color:#BAD0EA}}
.c-hl{{border-color:#FADBD8!important}}.c-ok{{border-color:#D5F5E3!important}}
.badge{{position:absolute;top:7px;right:8px;font-size:9px;color:#bbb;background:#F7F7F7;border-radius:3px;padding:1px 5px;letter-spacing:.3px}}
.lbl{{font-size:9.5px;color:#888;text-transform:uppercase;letter-spacing:.6px;line-height:1.5}}
.v{{font-size:21px;font-weight:700;margin:3px 0 2px;line-height:1.2}}
.s{{font-size:10.5px;color:#aaa;line-height:1.4}}.def{{font-size:9.5px;color:#5590CC;margin-top:5px;font-style:italic;line-height:1.3}}
.cb{{color:#2980B9}}.cr{{color:#E74C3C}}.cg{{color:#27AE60}}.co{{color:#E67E22}}
.chart-row{{display:grid;grid-template-columns:1.6fr 1fr;gap:12px;padding:10px 22px 14px}}
.cbox{{background:#fff;border-radius:10px;padding:12px;box-shadow:0 1px 3px rgba(0,0,0,.07)}}
.cbox h3{{font-size:12px;color:#1F4E79;font-weight:600;margin-bottom:6px;padding-bottom:5px;border-bottom:1px solid #EEE}}
.foot{{text-align:center;padding:14px;color:#bbb;font-size:10.5px;background:#fff;margin-top:14px;border-top:1px solid #EEE}}
dialog{{border:none;border-radius:12px;padding:0;width:min(1560px,97vw);max-height:90vh;box-shadow:0 20px 60px rgba(0,0,0,.28);
       position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);margin:0}}
dialog::backdrop{{background:rgba(0,0,0,.45)}}
.dlg-head{{background:#1F4E79;color:#fff;padding:12px 20px;display:flex;justify-content:space-between;align-items:center;border-radius:12px 12px 0 0}}
.dlg-head h3{{font-size:14px;font-weight:700}}
.close-btn{{background:rgba(255,255,255,.18);border:none;color:#fff;width:27px;height:27px;border-radius:50%;font-size:15px;cursor:pointer;line-height:27px;text-align:center}}
.dlg-note{{background:#EEF5FF;padding:7px 20px;font-size:10.5px;color:#555;border-bottom:1px solid #E0EAF5}}
.dlg-body{{overflow:auto;max-height:calc(88vh - 110px)}}
.dlg-body table{{width:100%;border-collapse:collapse;font-size:11px}}
.dlg-body th{{background:#2980B9;color:#fff;padding:6px 8px;text-align:left;position:sticky;top:0;white-space:nowrap}}
.dlg-body td{{padding:5px 8px;border-bottom:1px solid #F0F0F0;white-space:nowrap}}
.dlg-body td a{{color:#2980B9;font-weight:600;text-decoration:none}}
.dlg-body td a:hover{{text-decoration:underline}}
.dlg-body td.num{{text-align:right;font-variant-numeric:tabular-nums}}
.dlg-body tr:hover td{{background:#EEF5FF}}
.dlg-empty{{padding:28px;text-align:center;color:#aaa;font-style:italic}}
.dlg-body tr.row-bad td{{background:#FFF0EE!important}}
.dlg-body tr.row-ok td{{background:#F0FFF4!important}}
.dlg-body tr.row-warn td{{background:#FFFBEE!important}}
.dlg-body tr.row-bad:hover td{{background:#FFE5E2!important}}
.dlg-body tr.row-ok:hover td{{background:#E6FFEE!important}}
.dlg-body td.neg{{color:#C0392B;font-weight:600}}
.dlg-body td.pos{{color:#1E8449;font-weight:600}}
.brk-wrap{{padding:4px 22px 14px}}
.brk{{background:#fff;border-radius:10px;padding:14px 18px;box-shadow:0 1px 3px rgba(0,0,0,.08);border:2px solid #DCE7F3}}
.brk h3{{font-size:12.5px;color:#1F4E79;font-weight:700;margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid #EEE}}
.brk table{{width:100%;border-collapse:collapse;font-size:12px}}
.brk td{{padding:5px 8px;border-bottom:1px dashed #EEE}}
.brk td.num{{text-align:right;font-variant-numeric:tabular-nums;font-weight:700;white-space:nowrap;width:170px}}
.brk td.obs{{color:#999;font-size:10.5px;text-align:right;width:290px}}
.brk tr.pos td{{color:#1E8449}}
.brk tr.neg td{{color:#C0392B}}
.brk tr.tot td{{border-top:2px solid #1F4E79;border-bottom:none;font-size:13px;font-weight:700;padding-top:8px}}
#upd-banner{{display:none;position:fixed;bottom:18px;right:18px;z-index:99;background:#1F4E79;color:#fff;
  padding:10px 16px;border-radius:10px;box-shadow:0 8px 24px rgba(0,0,0,.3);align-items:center;gap:12px;font-size:12px}}
#upd-banner button{{background:#27AE60;border:none;color:#fff;padding:6px 14px;border-radius:6px;font-size:12px;font-weight:700;cursor:pointer}}
</style></head>
<body>

<div class="hdr">
  <div><h1>📊 Painel de Devoluções — Náutica Refrigeração</h1>
  <div class="sub">Portfólio: {d_start} → {d_end} &nbsp;|&nbsp; {kpi['total_pedidos']:,} pedidos &nbsp;|&nbsp; Gerado em {now_str}</div></div>
  {val_badge}
</div>

<div class="fbar">
  <label>PERÍODO:</label>
  <button class="btn-r" onclick="setR(7)">7 dias</button>
  <button class="btn-r" onclick="setR(30)">30 dias</button>
  <button class="btn-r" onclick="setR(90)">3 meses</button>
  <button class="btn-r" onclick="setR(180)">6 meses</button>
  <button class="btn-r active" id="btn-all" onclick="setR(0)">Todo o período</button>
  <span style="color:rgba(255,255,255,.3);margin:0 6px">|</span>
  <label>DE</label><input type="date" id="inp-s" onchange="fAll()">
  <label>ATÉ</label><input type="date" id="inp-e" onchange="fAll()">
  <span class="fperiod" id="fp"></span>
</div>
<div id="no-data-warn" style="display:none;background:#FFF3CD;border-left:4px solid #E67E22;padding:9px 22px;font-size:11.5px;color:#7D4A00">
  ⚠ <b>Nenhum dado para o período selecionado.</b> &nbsp;O portfólio importado cobre <b>{d_start}</b> até <b>{d_end}</b>. Use o botão <b>Todo o período</b> para ver todos os dados.
</div>

<!-- SEÇÃO 1: MOVIMENTO -->
<div class="sec-hdr"><h2>📦 Movimento do Período</h2><hr></div>
<div class="sec-note">
  <b>Cancelamento</b> = pedido cancelado antes da entrega — nenhum produto enviado &nbsp;|&nbsp;
  <b>Devolução</b> = produto entregue e devolvido; ML cobre automaticamente via BPP (Buyer Protection) &nbsp;|&nbsp;
  <b>Reclamação</b> = disputa formal; ML arbitra com base em fotos, rastreio e laudos
</div>
<div class="cards">
  <div class="card" onclick="openM('reclamacoes')"><span class="badge">ver lista</span>
    <div class="lbl">Total de Pedidos</div><div class="v cb" id="k-ped">{kpi['total_pedidos']:,}</div>
    <div class="s">no período selecionado</div></div>
  <div class="card" onclick="openM('cancelamentos')"><span class="badge">ver lista</span>
    <div class="lbl">❌ Cancelamentos</div><div class="v co" id="k-can-c">{len(df_cancel_real):,}</div>
    <div class="s" id="k-can-v">−{brl(df_cancel_real["valor_reemb"].sum())} reembolsados ao comprador</div>
    <div class="s" id="k-can-a" style="color:#C0392B;font-weight:600">⚠ {int(df_cancel_real["anom"].sum()):,} com saldo residual (anomalia)</div>
    <div class="def">Pedido cancelado antes da entrega — fonte: base de pedidos ML</div></div>
  <div class="card c-hl" onclick="openM('devolucoes')"><span class="badge">ver lista</span>
    <div class="lbl">📦 Devoluções (BPP)</div><div class="v cr" id="k-dev-c">{len(df_devol):,}</div>
    <div class="s" id="k-dev-v">{brl(df_devol["mp_valor_f"].sum())} em transações</div>
    <div class="def">Produto devolvido — ML debita automaticamente</div></div>
  <div class="card c-hl" onclick="openM('reclamacoes')"><span class="badge">ver lista</span>
    <div class="lbl">⚠️ Reclamações (Mediações)</div><div class="v cr" id="k-rec-c">{len(df_recl):,}</div>
    <div class="s" id="k-rec-v">{brl(df_recl["mp_valor_f"].sum())} em disputa</div>
    <div class="def">Disputa formal — ML decide com base em provas</div></div>
</div>

<!-- SEÇÃO 2: IMPACTO FINANCEIRO -->
<div class="sec-hdr"><h2>💰 Impacto Financeiro das Devoluções e Reclamações</h2><hr></div>
<div class="sec-note">Mediações encerradas na base Neon (ml_devolucoes). Clique em qualquer card para detalhar.</div>
<div class="cards">
  <div class="card c-hl" onclick="openM('reclamacoes')"><span class="badge">ver lista</span>
    <div class="lbl">Valor Total das Devoluções</div><div class="v cb" id="k-perda">{brl(kpi["perda_bruta"])}</div>
    <div class="s">valor original dos pedidos com débito</div></div>
  <div class="card" onclick="openM('reclamacoes')"><span class="badge">ver lista</span>
    <div class="lbl">Frete do Prejuízo</div><div class="v cr" id="k-frete">−{brl(frete_prej)}</div>
    <div class="s">tarifa de envio dos pedidos com perda</div>
    <div class="def">Exclui o frete das vendas normais</div></div>
  <div class="card c-ok" onclick="openM('reclamacoes')"><span class="badge">ver lista</span>
    <div class="lbl">Recuperado pelo ML</div><div class="v cg" id="k-rec">+{brl(kpi["recuperado_ml"])}</div>
    <div class="s">o que a ML reembolsou ao vendedor</div></div>
  <div class="card" style="border:2px solid #1F4E79" onclick="openM('reclamacoes')"><span class="badge">ver lista</span>
    <div class="lbl">% Recuperado</div><div class="v" style="color:{pct_color};font-size:24px" id="k-pct">{kpi["pct_recuperado"]:.1f}%</div>
    <div class="s">cobertura média da ML sobre as perdas</div></div>
  <div class="card" onclick="openM('reclamacoes')"><span class="badge">ver lista</span>
    <div class="lbl">Taxa ML Não Devolvida</div><div class="v cr" id="k-taxa">−{brl(kpi["taxa_retida"])}</div>
    <div class="s">comissão perdida nas perdas reais</div>
    <div class="def">Nas revertidas o ML estorna a tarifa (conciliado)</div></div>
  <div class="card c-hl" style="border-color:#C0392B" onclick="openM('reclamacoes')"><span class="badge">ver lista</span>
    <div class="lbl">⚠ Perda Real do Vendedor</div><div class="v cr" id="k-liq">−{brl(kpi["perda_liquida"])}</div>
    <div class="s">o que você efetivamente perdeu</div></div>
  <div class="card c-ok" onclick="openM('revertidos')"><span class="badge">ver lista</span>
    <div class="lbl">Revertidos do Negativo</div><div class="v cg" id="k-rev">{len(df_revert):,}</div>
    <div class="s">ML cobriu ≥ 90% — saiu no zero</div>
    <div class="def">Devolução onde você não saiu no prejuízo</div></div>
</div>

<!-- COMPOSIÇÃO DO IMPACTO — cada recorte que forma o valor total -->
<div class="brk-wrap"><div class="brk">
  <h3>🧮 Composição do Impacto — recortes do valor total (acompanha o período selecionado)</h3>
  <table>
    <tr><td>Valor original das devoluções e reclamações</td><td class="num" id="b-tot">{brl(kpi["perda_bruta"])}</td><td class="obs">100% — soma dos pedidos em disputa</td></tr>
    <tr class="pos"><td>(+) Recuperado pelo ML</td><td class="num" id="b-rec">+{brl(kpi["recuperado_ml"])}</td><td class="obs" id="b-rec-p">{kpi["pct_recuperado"]:.1f}% do valor original</td></tr>
    <tr class="pos"><td>(+) Mantido pelo vendedor — disputa encerrada sem reembolso</td><td class="num" id="b-mant">+{brl(kpi.get("mantido_vendedor",0))}</td><td class="obs">pagamento 'accredited' — conciliado via ML API + relatório MP</td></tr>
    <tr class="neg"><td>(−) Perda real do vendedor (não recuperado)</td><td class="num" id="b-liq">−{brl(kpi["perda_liquida"])}</td><td class="obs" id="b-liq-p"></td></tr>
    <tr class="neg"><td>(−) Taxa ML não devolvida (comissão retida)</td><td class="num" id="b-taxa">−{brl(kpi["taxa_retida"])}</td><td class="obs">custo extra — fora do valor original</td></tr>
    <tr class="neg"><td>(−) Frete dos pedidos com prejuízo</td><td class="num" id="b-frete">−{brl(frete_prej)}</td><td class="obs">custo extra — fora do valor original</td></tr>
    <tr class="tot neg"><td>(=) IMPACTO LÍQUIDO NO CAIXA</td><td class="num" id="b-imp">−{brl(kpi["perda_liquida"]+kpi["taxa_retida"]+frete_prej)}</td><td class="obs">perda real + taxa retida + frete do prejuízo</td></tr>
  </table>
</div></div>

<!-- SEÇÃO 3: RELATÓRIO MP -->
<div class="sec-hdr"><h2>🏦 Relatório Mercado Pago — after_collection</h2><hr></div>
<div class="sec-note">Arquivo importado do MP. Período: {df_mp_c["dia"].min()} → {df_mp_c["dia"].max()}</div>
<div class="cards">
  <div class="card c-ok" onclick="openM('devolucoes')"><span class="badge">ver lista</span>
    <div class="lbl">Proteções ML — BPP (Devoluções Automáticas)</div><div class="v cg" id="k-bpp-c">{len(df_devol):,}</div>
    <div class="s" id="k-bpp-v">{brl(df_devol["mp_valor_f"].sum())} = ML arcou automaticamente</div>
    <div class="def">bpp_refunded + bpp_covered</div></div>
  <div class="card" onclick="openM('mediacao')"><span class="badge">ver lista</span>
    <div class="lbl">Mediações Conciliadas</div><div class="v cb" id="k-med-c">{med_c:,}</div>
    <div class="s" id="k-med-v">{brl(med_v)} = ML cobriu via mediação</div>
    <div class="def">reconciled + compensated</div></div>
  <div class="card c-hl" onclick="openM('nao_conciliado')"><span class="badge">ver lista</span>
    <div class="lbl">Não Conciliados — Perda</div><div class="v cr" id="k-nc-c">{nc_c:,}</div>
    <div class="s" id="k-nc-v">{brl(nc_v)} = perda confirmada</div>
    <div class="def">Reclamação que o vendedor perdeu</div></div>
  <div class="card" onclick="openM('cancelamentos')"><span class="badge">ver lista</span>
    <div class="lbl">Cancelamentos Diretos</div><div class="v co" id="k-ref-c">{len(df_cancel):,}</div>
    <div class="s" id="k-ref-v">{brl(df_cancel["mp_valor_f"].sum())}</div>
    <div class="def">refunded = vendedor reembolsou diretamente</div></div>
</div>

<!-- GRÁFICOS -->
<div class="chart-row">
  <div class="cbox"><h3>📈 Impacto por Período — Recuperado vs Perda Real (arraste o slider para filtrar)</h3>{c_bar}</div>
  <div class="cbox"><h3>🔵 Tipos de Resolução (Relatório MP)</h3>{c_pie}</div>
</div>

{modals}

<div id="upd-banner"><span>🔄 Novos dados da ML API disponíveis</span><button onclick="location.reload()">Atualizar painel</button></div>

<div class="foot">Painel gerado automaticamente — processar_relatorios_mp.py • Náutica Refrigeração • {now_str}</div>

<script>
const D={embedded};
let cS=D.d_start,cE=D.d_end;
const fmtR=v=>v==null?'—':'R$\xa0'+Number(v).toLocaleString('pt-BR',{{minimumFractionDigits:2}});
const fmtN=v=>Number(v||0).toLocaleString('pt-BR');
// prejuízo SEMPRE negativo e vermelho; positivo SEMPRE verde — sinal no próprio texto
const fmtNeg=v=>{{const a=Math.abs(Number(v||0));return a<0.005?'R$\xa00,00':'−R$\xa0'+a.toLocaleString('pt-BR',{{minimumFractionDigits:2}});}};
const fmtPos=v=>{{const a=Math.abs(Number(v||0));return a<0.005?'R$\xa00,00':'+R$\xa0'+a.toLocaleString('pt-BR',{{minimumFractionDigits:2}});}};
const fmtPct=v=>Number(v||0).toFixed(1).replace('.',',')+'%';
function inR(d){{if(!d||d==='None'||d==='nan')return false;const s=String(d).slice(0,10);return(!cS||s>=cS)&&(!cE||s<=cE);}}
const sum=(a,k)=>a.reduce((s,r)=>s+(r[k]||0),0);

function calc(){{
  const pt=D.daily_port.filter(r=>inR(r.dia));
  const dv=D.daily_dev.filter(r=>inR(r.dia));
  const dl=D.daily_devol.filter(r=>inR(r.dia));
  const rc=D.daily_recl.filter(r=>inR(r.dia));
  const cn=D.daily_cancel.filter(r=>inR(r.dia));
  const cr=(D.daily_cancel_real||[]).filter(r=>inR(r.dia));
  const p=sum(dv,'perda'),r=sum(dv,'recuperado'),t=sum(dv,'taxa'),l=sum(dv,'liq'),mt=sum(dv,'mantido');
  return{{ped:sum(pt,'pedidos'),
    can_c:sum(cr,'count'),can_v:sum(cr,'valor'),can_a:sum(cr,'anom'),
    dev_c:sum(dl,'count'),dev_v:sum(dl,'valor'),
    rec_c:sum(rc,'count'),rec_v:sum(rc,'valor'),
    perda:p,recup:r,taxa:t,liq:l,mant:mt,pct:p>0?r/p*100:0,
    rev:D.revert_modal.filter(r=>inR(r.dia)).length,
    bpp_c:sum(dl,'count'),bpp_v:sum(dl,'valor'),
    med_c:D.mediacao_modal.filter(r=>inR(r.dia)).length,
    med_v:D.mediacao_modal.filter(r=>inR(r.dia)).reduce((s,r)=>s+(r.mp_valor_f||0),0),
    nc_c:D.nc_modal.filter(r=>inR(r.dia)).length,
    nc_v:D.nc_modal.filter(r=>inR(r.dia)).reduce((s,r)=>s+(r.mp_valor_f||0),0),
    ref_c:sum(cn,'count'),ref_v:sum(cn,'valor')}};
}}

function set(id,v){{const e=document.getElementById(id);if(e)e.textContent=v;}}
function setC(id,pct){{const e=document.getElementById(id);if(e)e.style.color=pct>=70?'#27AE60':pct>=40?'#E67E22':'#E74C3C';}}

function upCards(){{
  const k=calc();
  // Frete dinâmico: apenas pedidos com prejuízo, filtrado pelo período
  const fpSum=(D.daily_frete_p||[]).filter(r=>inR(r.dia)).reduce((s,r)=>s+(r.frete_prej||0),0);
  set('k-frete', fmtNeg(fpSum));
  // Aviso quando período não tem dados
  const warn=document.getElementById('no-data-warn');
  if(warn) warn.style.display=(k.ped===0&&k.dev_c===0&&k.can_c===0&&k.rec_c===0)?'block':'none';
  set('k-ped',fmtN(k.ped));
  set('k-can-c',fmtN(k.can_c));set('k-can-v',fmtNeg(k.can_v)+' reembolsados ao comprador');
  const ka=document.getElementById('k-can-a');
  if(ka){{ka.textContent='⚠ '+fmtN(k.can_a)+' com saldo residual (anomalia)';ka.style.display=k.can_a>0?'block':'none';}}
  set('k-dev-c',fmtN(k.dev_c));set('k-dev-v',fmtR(k.dev_v)+' em transações');
  set('k-rec-c',fmtN(k.rec_c));set('k-rec-v',fmtR(k.rec_v)+' em disputa');
  set('k-perda',fmtR(k.perda));set('k-rec',fmtPos(k.recup));
  set('k-taxa',fmtNeg(k.taxa));set('k-liq',fmtNeg(k.liq));
  set('k-pct',fmtPct(k.pct));setC('k-pct',k.pct);
  set('k-rev',fmtN(k.rev));
  set('k-bpp-c',fmtN(k.bpp_c));set('k-bpp-v',fmtPos(k.bpp_v)+' = ML arcou automaticamente');
  set('k-med-c',fmtN(k.med_c));set('k-med-v',fmtPos(k.med_v)+' = ML cobriu via mediação');
  set('k-nc-c',fmtN(k.nc_c));set('k-nc-v',fmtNeg(k.nc_v)+' = perda confirmada');
  set('k-ref-c',fmtN(k.ref_c));set('k-ref-v',fmtR(k.ref_v));
  // Breakdown — recortes que compõem o valor total (sempre sincronizado com o filtro)
  const imp=k.liq+k.taxa+fpSum;
  set('b-tot',fmtR(k.perda));
  set('b-rec',fmtPos(k.recup));set('b-rec-p',fmtPct(k.perda>0?k.recup/k.perda*100:0)+' do valor original');
  set('b-mant',fmtPos(k.mant));
  set('b-liq',fmtNeg(k.liq)); set('b-liq-p',fmtPct(k.perda>0?k.liq/k.perda*100:0)+' do valor original');
  set('b-taxa',fmtNeg(k.taxa));set('b-frete',fmtNeg(fpSum));
  set('b-imp',fmtNeg(imp));
  set('fp','📅 '+cS+' → '+cE+' ('+fmtN(k.ped)+' pedidos)');
}}

function upCharts(){{
  if(typeof Plotly==='undefined')return;
  document.querySelectorAll('.js-plotly-plot').forEach(el=>{{try{{Plotly.relayout(el,{{'xaxis.range':[cS,cE]}});}}catch(e){{}}}}); 
}}

function persistR(){{try{{sessionStorage.setItem('pd_range',cS+'|'+cE);}}catch(e){{}}}}

function fAll(){{cS=document.getElementById('inp-s').value||D.d_start;cE=document.getElementById('inp-e').value||D.d_end;document.querySelectorAll('.btn-r').forEach(b=>b.classList.remove('active'));persistR();upCards();upCharts();}}

function setR(days){{
  document.querySelectorAll('.btn-r').forEach(b=>b.classList.remove('active'));
  if(days===0){{document.getElementById('btn-all')?.classList.add('active');cS=D.d_start;cE=D.d_end;}}
  else{{const e=new Date(D.d_end),s=new Date(+e-days*86400000);cS=s.toISOString().slice(0,10);cE=D.d_end;}}  // ref=fim dos dados, não hoje
  document.getElementById('inp-s').value=cS;
  document.getElementById('inp-e').value=cE;
  persistR();upCards();upCharts();
}}

// tipos de célula: 0=texto | 1=número neutro | 2=prejuízo (−, vermelho) | 3=positivo (+, verde)
const COLS={{
  reclamacoes:    [['Nº Pedido','order_id',0],['Data','dia',0],['Produto','sku',0],['Motivo da Reclamação','motivo_ml',0],['Val. Original R$','perda_bruta',1],['Recuperado ML R$','recuperado_ml',3],['Taxa ML R$','taxa_ml_retida',2],['Perda Real R$','perda_liquida',2],['Situação','situacao',0],['Estado Atual ML','estado_api',0],['Validado em','api_em',0]],
  devolucoes:     [['Nº Pedido','mp_order_id',0],['Data','dia',0],['Tipo de Resolução','status_detail',0],['Valor R$','mp_valor_f',3],['Situação','situacao',0],['Estado Atual ML','estado_api',0],['Validado em','api_em',0]],
  cancelamentos:  [['Nº Pedido','order_id',0],['Data','dia',0],['Quem Cancelou','tipo',0],['Motivo','motivo',0],['Reembolsado R$','valor_reemb',2],['Saldo Residual R$','resid',2],['⚠ Situação','situacao',0],['Estado Atual ML','estado_api',0],['Validado em','api_em',0]],
  mediacao:       [['Nº Pedido','mp_order_id',0],['Data','dia',0],['Tipo de Resolução','status_detail',0],['Valor R$','mp_valor_f',3],['Situação','situacao',0],['Estado Atual ML','estado_api',0],['Validado em','api_em',0]],
  nao_conciliado: [['Nº Pedido','mp_order_id',0],['Data','dia',0],['Tipo','status_detail',0],['Valor R$','mp_valor_f',2],['Situação','situacao',0],['Estado Atual ML','estado_api',0],['Validado em','api_em',0]],
  revertidos:     [['Nº Pedido','order_id',0],['Data','dia',0],['Produto','sku',0],['Val. Original R$','perda_bruta',1],['Recuperado R$','recuperado_ml',3],['Situação','situacao',0],['Estado Atual ML','estado_api',0]],
}};
const DM={{reclamacoes:'recl_modal',devolucoes:'devol_modal',cancelamentos:'cancel_modal',mediacao:'mediacao_modal',nao_conciliado:'nc_modal',revertidos:'revert_modal'}};

function openM(id){{const d=document.getElementById('dlg-'+id);if(d){{buildT(id);d.showModal();}}}}

function buildT(id){{
  const tb=document.getElementById('tbody-'+id);if(!tb)return;
  const rows=(D[DM[id]]||D.recl_modal).filter(r=>inR(r.dia));
  const cols=COLS[id]||COLS.reclamacoes;
  const cnt=document.getElementById('cnt-'+id);if(cnt)cnt.textContent=rows.length.toLocaleString('pt-BR');
  tb.innerHTML='';
  if(!rows.length){{tb.innerHTML='<tr><td colspan="10" class="dlg-empty">Nenhum registro para o período selecionado.</td></tr>';return;}}
  rows.slice(0,500).forEach(r=>{{
    const tr=document.createElement('tr');
    if(r.classe) tr.className='row-'+r.classe;
    cols.forEach(([l,k,t])=>{{
      const td=document.createElement('td');
      const v=r[k];
      if((k==='order_id'||k==='mp_order_id')&&v!=null){{
        // pedido clicável → abre a venda na plataforma Meli em nova aba
        const a=document.createElement('a');
        a.href='https://www.mercadolivre.com.br/vendas/'+String(v)+'/detalhe';
        a.target='_blank';a.rel='noopener';a.title='Abrir esta venda no Mercado Livre';
        a.textContent=String(v);
        td.appendChild(a);
      }}
      else if(t===1){{td.textContent=v!=null?Number(v).toLocaleString('pt-BR',{{minimumFractionDigits:2}}):'0,00';td.className='num';}}
      else if(t===2){{const a=Math.abs(Number(v||0));
        td.textContent=a<0.005?'0,00':'−'+a.toLocaleString('pt-BR',{{minimumFractionDigits:2}});
        td.className='num'+(a<0.005?'':' neg');}}
      else if(t===3){{const a=Math.abs(Number(v||0));
        td.textContent=a<0.005?'0,00':'+'+a.toLocaleString('pt-BR',{{minimumFractionDigits:2}});
        td.className='num'+(a<0.005?'':' pos');}}
      else td.textContent=v!=null?String(v):'—';
      tr.appendChild(td);
    }});
    tb.appendChild(tr);
  }});
}}

document.addEventListener('DOMContentLoaded',()=>{{
  // Restaura o período escolhido antes de um reload (atualização automática)
  try{{
    const sv=sessionStorage.getItem('pd_range');
    if(sv){{const[a,b]=sv.split('|');if(a>=D.d_start&&b<=D.d_end&&a<=b){{cS=a;cE=b;}}}}
  }}catch(e){{}}
  document.getElementById('inp-s').value=cS;
  document.getElementById('inp-e').value=cE;
  upCards();
}});

// ── Atualização automática (polling) ─────────────────────────────────────────
// Compara o 'gerado_em' do JSON gêmeo a cada 60s. Quando o ml_live_poll.py
// regenera o painel, aparece o banner para recarregar (filtro é preservado).
// Em file:// o fetch é bloqueado pelo navegador — o banner simplesmente não
// aparece; use "python scripts/ml_live_poll.py --serve 8765" para o modo vivo.
(function(){{
  if(location.protocol==='file:')return;
  const jsonUrl=location.pathname.replace(/\\.html?$/i,'.json');
  if(jsonUrl===location.pathname)return;
  let first=null;
  setInterval(async()=>{{
    try{{
      const r=await fetch(jsonUrl+'?t='+Date.now(),{{cache:'no-store'}});
      if(!r.ok)return;
      const j=await r.json();
      const g=j.gerado_em||'';
      if(first===null){{first=g;return;}}
      if(g&&g!==first){{const b=document.getElementById('upd-banner');if(b)b.style.display='flex';}}
    }}catch(e){{}}
  }},60000);
}})();
</script>
</body></html>"""

    output_path.write_text(html, encoding="utf-8")
    print(f"  ✓ Dashboard HTML salvo em: {output_path}")




def _build_xlsx(df_mp: pd.DataFrame, df_port: pd.DataFrame, kpi: dict, output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        from openpyxl.chart.series import DataPoint
    except ImportError:
        print("  ⚠ openpyxl não instalado – XLSX ignorado")
        return

    wb = Workbook()
    wb.remove(wb.active)

    DARK   = PatternFill("solid", fgColor="1F4E79")
    MED    = PatternFill("solid", fgColor="2E75B6")
    LIGHT  = PatternFill("solid", fgColor="D6E4F0")
    GOOD   = PatternFill("solid", fgColor="C6EFCE")
    BAD    = PatternFill("solid", fgColor="FFC7CE")
    WARN   = PatternFill("solid", fgColor="FFEB9C")
    WHITE  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    NORMAL = Font(name="Calibri", size=10)

    def _hdr(ws, row: int, fill=DARK):
        for cell in ws[row]:
            if cell.value is not None:
                cell.fill = fill
                cell.font = WHITE
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

    def _auto_w(ws, extra=4, mx=55):
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + extra, mx)

    def _write_df(ws, df: pd.DataFrame, sr=1, sc=1):
        for ci, col in enumerate(df.columns, sc):
            ws.cell(sr, ci, str(col))
        for ri, row_data in enumerate(df.itertuples(index=False), sr + 1):
            for ci, val in enumerate(row_data, sc):
                if isinstance(val, float) and np.isnan(val):
                    val = None
                elif isinstance(val, pd.Timestamp):
                    val = val.replace(tzinfo=None)  # openpyxl não suporta tz-aware
                elif hasattr(val, "tzinfo") and getattr(val, "tzinfo", None) is not None:
                    val = val.replace(tzinfo=None)
                ws.cell(ri, ci, val)

    # ── Aba 1: Resumo Geral ───────────────────────────────────────────────────
    ws1 = wb.create_sheet("Resumo Geral")
    ws1["A1"] = "ANÁLISE FINANCEIRA DE DEVOLUÇÕES – NAUTICA REFRIGERAÇÃO"
    ws1["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
    ws1["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1["A2"].font = Font(italic=True, color="666666", name="Calibri")
    ws1.merge_cells("A1:C1")
    ws1.merge_cells("A2:C2")

    kpi_rows = [
        ("PORTFÓLIO COMPLETO", None, None),
        ("Total de pedidos",               kpi["total_pedidos"],    "abr/2025 → mai/2026"),
        ("Taxa ML (tarifas + impostos)",    kpi["taxa_ml"],          "custo da plataforma Mercado Livre"),
        ("Tarifa de frete (saída)",         kpi["tarifa_frete"],     "custo logístico cobrado pela ML"),
        ("Frete cobrado do comprador",      kpi["frete_cobrado"],    "receita de envio recebida"),
        ("Saldo frete (cobrado − tarifa)",  kpi["saldo_frete"],      "positivo = comprador cobre o frete"),
        ("Cancelamentos/Reembolsos",        kpi["cancelamentos"],    "reembolsos saídos da conta"),
        ("Receita líquida (total_brl)",     kpi["receita_liquida"],  "após tarifas e cancelamentos"),
        ("DEVOLUÇÕES E MEDIAÇÕES ML", None, None),
        ("Pedidos com devolução/mediação",  kpi["dev_pedidos"],      f"{PCT(kpi['dev_pct_portfolio'])} do portfólio"),
        ("Perda bruta total",               kpi["perda_bruta"],      "valor original dos pedidos devolvidos"),
        ("Valor recuperado pelo ML",        kpi["recuperado_ml"],    "ML reembolsou ao vendedor"),
        ("Mantido pelo vendedor",           kpi.get("mantido_vendedor", 0), "disputa encerrada sem reembolso (accredited)"),
        ("Taxa ML retida (não devolvida)",  kpi["taxa_retida"],      "taxa perdida nas perdas reais"),
        ("⚠ Perda líquida real",           kpi["perda_liquida"],    "o que efetivamente você perdeu"),
        ("% médio recuperado",             PCT(kpi["pct_recuperado"]), "quanto a ML cobriu em média"),
        ("RELATÓRIO MP INGERIDO", None, None),
        ("Transações MP processadas",       kpi["mp_total_tx"],      "linhas importadas no Neon"),
        ("Pedidos únicos no MP",            kpi["mp_pedidos_unicos"], None),
        ("Valor total das transações MP",   kpi["mp_valor_total"],   "soma de todas as transações"),
    ]

    linha = 4
    for hdr in [("INDICADOR", "VALOR", "OBSERVAÇÃO")]:
        for ci, h in enumerate(hdr, 1):
            c = ws1.cell(linha, ci, h)
            c.fill = DARK; c.font = WHITE
            c.alignment = Alignment(horizontal="center")
    linha += 1

    for ind, val, obs in kpi_rows:
        c1 = ws1.cell(linha, 1, ind)
        c2 = ws1.cell(linha, 2)
        c3 = ws1.cell(linha, 3, obs or "")
        if val is None:
            c1.fill = MED; c1.font = WHITE
            c2.fill = MED; c3.fill = MED
        else:
            c1.fill = LIGHT
            if isinstance(val, (int, float)) and not isinstance(val, bool) and val > 100:
                c2.value = float(val)
                c2.number_format = 'R$ #,##0.00'
            else:
                c2.value = val
        linha += 1

    ws1.column_dimensions["A"].width = 45
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 45

    # ── Aba 2: Por Mês ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por Mes")
    ws2["A1"] = "ANÁLISE MENSAL"
    ws2["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Calibri")

    df_dev = df_port[df_port["perda_bruta"] > 0].drop_duplicates("order_id")
    mes_port = (
        df_port.drop_duplicates("order_id")
        .groupby("mes")
        .agg(
            pedidos=("order_id","count"),
            receita_bruta=("receita_bruta","sum"),
            taxa_ml=("taxa_ml","sum"),
            tarifa_frete=("tarifa_frete","sum"),
            cancelamentos=("cancelamentos","sum"),
            receita_liquida=("total_brl","sum"),
        ).reset_index()
    )
    mes_dev = (
        df_dev.groupby("mes")
        .agg(
            dev_pedidos=("order_id","count"),
            perda_bruta=("perda_bruta","sum"),
            recuperado=("recuperado_ml","sum"),
            taxa_retida=("taxa_ml_retida","sum"),
            perda_liquida=("perda_liquida","sum"),
        ).reset_index()
    )
    mes_fin = mes_port.merge(mes_dev, on="mes", how="left").sort_values("mes")
    mes_fin["pct_rec"] = (mes_fin["recuperado"] / mes_fin["perda_bruta"].where(mes_fin["perda_bruta"]>0)*100).round(1)
    mes_fin.rename(columns={
        "mes":"Mês","pedidos":"Pedidos","receita_bruta":"Receita Bruta (R$)",
        "taxa_ml":"Custo ML — Tarifas (R$)","tarifa_frete":"Custo de Envio ML (R$)",
        "cancelamentos":"Cancelamentos e Reembolsos (R$)","receita_liquida":"Receita Líquida (R$)",
        "dev_pedidos":"Pedidos c/ Devolução","perda_bruta":"Valor Total Devolvido (R$)",
        "recuperado":"Recuperado pelo ML (R$)","taxa_retida":"Taxa ML Não Devolvida (R$)",
        "perda_liquida":"Perda Real do Vendedor (R$)","pct_rec":"% Recuperado pelo ML",
    }, inplace=True)
    _write_df(ws2, mes_fin, sr=3)
    _hdr(ws2, 3)
    ws2.freeze_panes = "A4"
    _auto_w(ws2)

    # Gráfico de barras embutido (Per Mês – Perda vs Recuperado)
    if len(mes_fin) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.title = "Perda Bruta vs Recuperado por Mês (R$)"
        chart.y_axis.title = "R$"
        chart.x_axis.title = "Mês"
        chart.width  = 22
        chart.height = 14

        # Achar colunas
        cols_list = list(mes_fin.columns)
        nrows = len(mes_fin) + 1

        def _col_idx(name: str) -> int | None:
            try:
                return cols_list.index(name) + 1
            except ValueError:
                return None

        for col_name, label in [("Recuperado ML R$","Recuperado"), ("Perda Líquida R$","Perda Líquida"), ("Taxa Retida R$","Taxa Retida")]:
            ci = _col_idx(col_name)
            if ci:
                data = Reference(ws2, min_col=ci, max_col=ci, min_row=3, max_row=3+len(mes_fin))
                chart.add_data(data, titles_from_data=True)

        cat = Reference(ws2, min_col=1, min_row=4, max_row=3+len(mes_fin))
        chart.set_categories(cat)
        ws2.add_chart(chart, "P3")

    # ── Aba 3: Por Semestre ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Análise Semestral")
    ws3["A1"] = "ANÁLISE SEMESTRAL — RECEITA, CUSTOS E DEVOLUÇÕES"
    ws3["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Calibri")
    sem = (
        df_port.drop_duplicates("order_id")
        .groupby("semestre")
        .agg(pedidos=("order_id","count"), receita_bruta=("receita_bruta","sum"),
             taxa_ml=("taxa_ml","sum"), tarifa_frete=("tarifa_frete","sum"),
             cancelamentos=("cancelamentos","sum"), receita_liquida=("total_brl","sum"))
        .reset_index()
        .rename(columns={"semestre":"Semestre","pedidos":"Pedidos",
                          "receita_bruta":"Receita Bruta (R$)","taxa_ml":"Custo ML Tarifas (R$)",
                          "tarifa_frete":"Custo de Envio ML (R$)","cancelamentos":"Cancelamentos (R$)",
                          "receita_liquida":"Receita Líquida (R$)"})
    )
    _write_df(ws3, sem, sr=3)
    _hdr(ws3, 3)
    _auto_w(ws3)

    # ── Aba 4: Por Produto ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Produtos com Devolução")
    ws4["A1"] = "PRODUTOS COM MAIOR IMPACTO EM DEVOLUÇÕES — TOP 60 (ordenado por valor devolvido)"
    ws4["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Calibri")
    sku_agg = (
        df_dev.groupby(["sku"])
        .agg(ocorrencias=("order_id","count"), perda_bruta=("perda_bruta","sum"),
             recuperado=("recuperado_ml","sum"), taxa_retida=("taxa_ml_retida","sum"),
             perda_liquida=("perda_liquida","sum"), cmv_total=("cmv","sum"))
        .reset_index()
        .sort_values("perda_bruta", ascending=False)
        .head(60)
        .rename(columns={"sku":"Cód. do Produto (SKU)","ocorrencias":"Qtd. de Devoluções",
                          "perda_bruta":"Valor Total Devolvido (R$)","recuperado":"Recuperado pelo ML (R$)",
                          "taxa_retida":"Taxa ML Não Devolvida (R$)","perda_liquida":"Perda Real do Vendedor (R$)",
                          "cmv_total":"Custo Total (CMV) (R$)"})
    )
    sku_agg["% Recuperado pelo ML"] = (sku_agg["Recuperado pelo ML (R$)"]/sku_agg["Valor Total Devolvido (R$)"].where(sku_agg["Valor Total Devolvido (R$)"]>0)*100).round(1)
    _write_df(ws4, sku_agg, sr=3)
    _hdr(ws4, 3)
    ws4.freeze_panes = "A4"
    _auto_w(ws4)

    # ── Aba 5: Motivos ────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Motivos das Reclamações")
    ws5["A1"] = "DEVOLUÇÕES POR MOTIVO DE RECLAMAÇÃO (fonte: Mercado Livre)"
    ws5["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Calibri")
    mot = (
        df_dev.groupby("motivo_ml")
        .agg(ocorrencias=("order_id","count"), perda_bruta=("perda_bruta","sum"),
             recuperado=("recuperado_ml","sum"), taxa_retida=("taxa_ml_retida","sum"),
             perda_liquida=("perda_liquida","sum"))
        .reset_index()
        .sort_values("perda_bruta", ascending=False)
        .rename(columns={"motivo_ml":"Motivo da Reclamação","ocorrencias":"Qtd. de Ocorrências",
                          "perda_bruta":"Valor Total Devolvido (R$)","recuperado":"Recuperado pelo ML (R$)",
                          "taxa_retida":"Taxa ML Não Devolvida (R$)","perda_liquida":"Perda Real do Vendedor (R$)"})
    )
    mot["% Recuperado pelo ML"] = (mot["Recuperado pelo ML (R$)"]/mot["Valor Total Devolvido (R$)"].where(mot["Valor Total Devolvido (R$)"]>0)*100).round(1)
    _write_df(ws5, mot, sr=3)
    _hdr(ws5, 3)
    ws5.freeze_panes = "A4"
    _auto_w(ws5)

    # ── Aba 6: Relatório MP ───────────────────────────────────────────────────
    ws6 = wb.create_sheet("Relatório Mercado Pago")
    ws6["A1"] = "ANÁLISE DO RELATÓRIO MERCADO PAGO (after_collection)"
    ws6["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Calibri")
    mp_buck = (
        df_mp.dropna(subset=["status_detail"])
        .groupby(["status_detail","categoria"])
        .agg(transacoes=("mp_order_id","count"), pedidos_unicos=("mp_order_id","nunique"),
             valor_total=("mp_valor","sum"))
        .reset_index()
        .sort_values("valor_total", ascending=False)
    )
    mp_buck["tipo_resolucao_pt"] = mp_buck["status_detail"].map(lambda s: STATUS_DETAIL_LABEL.get(str(s), str(s)))
    mp_buck["pct_do_total"] = (mp_buck["valor_total"]/mp_buck["valor_total"].sum()*100).round(1)
    mp_buck = mp_buck.rename(columns={
        "tipo_resolucao_pt":"Tipo de Resolução","categoria":"Categoria de Impacto",
        "transacoes":"Nº de Transações","pedidos_unicos":"Pedidos Únicos",
        "valor_total":"Valor Total (R$)","pct_do_total":"% do Total",
    })[["Tipo de Resolução","Categoria de Impacto","Nº de Transações","Pedidos Únicos","Valor Total (R$)","% do Total"]]
    _write_df(ws6, mp_buck, sr=3)
    _hdr(ws6, 3)
    cat_fills = {
        "Protegido ML": GOOD, "Mediação ML": PatternFill("solid", fgColor="DDEBF7"),
        "Perda Confirmada": BAD, "Reembolso Direto": WARN,
        "Administrativo": PatternFill("solid", fgColor="E2EFDA"),
    }
    for ri in range(4, 4 + len(mp_buck)):
        cv = str(ws6.cell(ri, 2).value or "")
        fill = cat_fills.get(cv)
        if fill:
            for ci in range(1, len(mp_buck.columns) + 1):
                ws6.cell(ri, ci).fill = fill
    _auto_w(ws6)

    # ── Aba 7: Mapa Completo Devoluções ───────────────────────────────────────
    ws7 = wb.create_sheet("Detalhe por Pedido")
    ws7["A1"] = f"DETALHE COMPLETO POR PEDIDO — {len(df_dev):,} DEVOLUÇÕES"
    ws7["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Calibri")
    mapa = df_dev[[
        "order_id","mes","sku","motivo_ml",
        "perda_bruta","taxa_ml_retida","recuperado_ml","perda_liquida","impacto_cmv","cmv"
    ]].copy().rename(columns={
        "order_id":"Nº do Pedido","mes":"Mês da Devolução",
        "sku":"Cód. do Produto (SKU)","motivo_ml":"Motivo da Reclamação (ML)",
        "perda_bruta":"Valor Original do Pedido (R$)",
        "taxa_ml_retida":"Taxa ML Não Devolvida (R$)",
        "recuperado_ml":"Recuperado pelo ML (R$)",
        "perda_liquida":"Perda Real do Vendedor (R$)",
        "impacto_cmv":"Impacto no Custo Total (R$)",
        "cmv":"Custo do Produto — CMV (R$)",
    }).sort_values("Perda Real do Vendedor (R$)", ascending=False)
    _write_df(ws7, mapa, sr=3)
    _hdr(ws7, 3)
    ws7.freeze_panes = "A4"
    _auto_w(ws7)

    # ── Aba 8: Log de Importações ─────────────────────────────────────────────
    ws8 = wb.create_sheet("Histórico de Importações")
    ws8["A1"] = "HISTÓRICO DE ARQUIVOS MERCADO PAGO IMPORTADOS"
    ws8["A1"].font = Font(bold=True, size=13, color="1F4E79", name="Calibri")
    conn = get_db_connection()
    try:
        from src.services.mp_ingestion import _ensure_tables
        _ensure_tables(conn)
        df_log = pd.read_sql(
            "SELECT filename, imported_at, rows_total, rows_new, rows_dup, periodo_inicio, periodo_fim "
            "FROM mp_import_log ORDER BY imported_at DESC", conn
        )
        df_log.rename(columns={
            "filename":"Arquivo Importado","imported_at":"Data de Importação",
            "rows_total":"Total de Linhas","rows_new":"Novas","rows_dup":"Duplicadas",
            "periodo_inicio":"Início do Período","periodo_fim":"Fim do Período",
        }, inplace=True)
        _write_df(ws8, df_log, sr=3)
        _hdr(ws8, 3)
        _auto_w(ws8)
    except Exception:
        ws8["A3"] = "Nenhum log disponível"
    finally:
        conn.close()

    wb.save(str(output_path))
    print(f"  ✓ XLSX salvo em: {output_path}")


# ─── Conciliação completa ML API + Tiny ─────────────────────────────────────

_VAL_DDL = """
CREATE TABLE IF NOT EXISTS mp_validation_results (
    order_id         BIGINT      PRIMARY KEY,
    validated_at     TIMESTAMPTZ NOT NULL DEFAULT NOW(),
    db_total         NUMERIC(14,2),
    api_total        NUMERIC(14,2),
    delta            NUMERIC(14,2),
    api_pay_status   TEXT,
    api_pay_detail   TEXT,
    concorda_ml      BOOLEAN,
    mp_status_detail TEXT,
    mp_valor         NUMERIC(14,2),
    sku              TEXT,
    tiny_cmv         NUMERIC(14,2),
    tiny_nome        TEXT
)
"""

def _ensure_validation_table(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(_VAL_DDL)
        cur.execute("CREATE INDEX IF NOT EXISTS idx_val_concorda ON mp_validation_results (concorda_ml)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_val_validated_at ON mp_validation_results (validated_at)")
    conn.commit()


# Estados finais na ML API — pedidos nesses estados não mudam mais; não vale a
# pena re-consultar no polling. Tudo fora daqui (pending, in_mediation, None,
# erros de rede…) ainda pode mudar e entra no ciclo incremental.
TERMINAL_PAY_DETAILS: frozenset[str] = frozenset({
    "bpp_refunded", "bpp_covered", "partially_bpp_refunded", "partially_bpp_covered",
    "refunded", "partially_refunded", "reconciled", "compensated",
    "by_admin", "by_payer", "accredited",
})


def _select_polling_ids(conn, janela_dias: int = 45, ttl_min: int = 30) -> set[int]:
    """Seleciona o conjunto MÍNIMO de pedidos que vale revalidar na ML API:

      (a) claims ainda abertos em ml_devolucoes (podem virar perda/reversão)
      (b) pedidos cancelados recentes (janela_dias) — saldo pode mudar
      (c) pedidos cujo último estado na API não é terminal (pending, erro, None)

    Exclui qualquer pedido já validado nos últimos `ttl_min` minutos, para o
    ciclo ser barato mesmo rodando em intervalo curto.
    """
    ids: set[int] = set()
    with conn.cursor() as cur:
        # (a) claims abertos
        cur.execute("""
            SELECT DISTINCT order_id::bigint FROM ml_devolucoes
            WHERE claim_status = 'opened' AND order_id IS NOT NULL
        """)
        ids |= {r[0] for r in cur.fetchall()}

        # (b) cancelamentos recentes
        cur.execute("""
            SELECT DISTINCT order_id::bigint FROM orders
            WHERE (estado ILIKE 'Cancelada%%' OR estado ILIKE 'Pacote cancelado%%'
                   OR estado ILIKE 'Venda cancelada%%' OR estado ILIKE 'Você cancelou%%')
              AND data_venda >= NOW() - make_interval(days => %s)
        """, (janela_dias,))
        ids |= {r[0] for r in cur.fetchall()}

        # (c) estados não-terminais no cache de validação
        cur.execute("""
            SELECT order_id FROM mp_validation_results
            WHERE api_pay_detail IS NULL
               OR (api_pay_detail NOT LIKE 'cc_rejected%%'
                   AND api_pay_detail <> ALL(%s))
        """, (list(TERMINAL_PAY_DETAILS),))
        ids |= {r[0] for r in cur.fetchall()}

        # tira os validados há pouco (TTL)
        cur.execute("""
            SELECT order_id FROM mp_validation_results
            WHERE validated_at >= NOW() - make_interval(mins => %s)
        """, (ttl_min,))
        recentes = {r[0] for r in cur.fetchall()}

    return ids - recentes


def _run_full_validation(conn, force: bool = False, n_workers: int = 8, validar_base_completa: bool = False, only_ids: set[int] | None = None) -> dict:
    """Valida TODA a população do relatório MP (e opcionalmente a base ml_devolucoes completa)
    contra a ML API. Resultados cacheados em mp_validation_results no Neon.

    Args:
        force: re-valida todos, ignorando o cache
        n_workers: threads paralelas para chamadas à ML API
        validar_base_completa: se True, inclui TODA a base ml_devolucoes (13+ meses, ~10 min)
        only_ids: se fornecido, revalida EXATAMENTE esses pedidos (modo polling
                  incremental — ignora cache para eles, não toca no resto)
    """
    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed

    _ensure_validation_table(conn)

    # 1. Todos os order_ids do relatório MP + seus metadados
    with conn.cursor() as cur:
        cur.execute("""
            SELECT order_id, status_detail, valor FROM mp_transactions
            WHERE order_id IS NOT NULL
        """)
        mp_rows: dict[int, dict] = {
            r[0]: {"status_detail": r[1], "mp_valor": float(r[2] or 0)}
            for r in cur.fetchall()
        }
    all_ids = list(mp_rows.keys())

    # Incluir base completa ml_devolucoes (13+ meses) se solicitado
    if validar_base_completa:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT DISTINCT order_id::bigint FROM ml_devolucoes
                WHERE claim_type = 'mediations' AND claim_status = 'closed'
                  AND order_id IS NOT NULL
            """)
            ml_ids = {r[0] for r in cur.fetchall()}
        novos_da_base = ml_ids - set(all_ids)
        for oid in novos_da_base:
            mp_rows[oid] = {"status_detail": "base_ml_devolucoes", "mp_valor": 0.0}
        all_ids = list(mp_rows.keys())
        print(f"  + Base ml_devolucoes: {len(novos_da_base):,} pedidos adicionais inclusos")

    # 2. Selecionar o que validar
    if only_ids is not None:
        # modo polling incremental: revalida exatamente estes (força re-consulta)
        for oid in only_ids:
            mp_rows.setdefault(oid, {"status_detail": "poll_incremental", "mp_valor": 0.0})
        all_ids = list(mp_rows.keys())
        todo = sorted(only_ids)
        print(f"  → polling incremental: {len(todo):,} pedidos a revalidar na ML API")
    elif not force:
        with conn.cursor() as cur:
            cur.execute("SELECT order_id FROM mp_validation_results")
            already = {r[0] for r in cur.fetchall()}
        todo = [oid for oid in all_ids if oid not in already]
        print(f"  → {len(all_ids):,} pedidos MP  |  {len(all_ids)-len(todo):,} já no cache  |  {len(todo):,} a validar agora")
    else:
        todo = all_ids
        print(f"  → {len(all_ids):,} pedidos MP  |  revalidação forçada de todos")

    # 3. Dados do DB para os pedidos a validar
    db_map: dict[int, dict] = {}
    tiny_map: dict[str, tuple] = {}
    if todo:
        try:
            df_db = pd.read_sql("""
                SELECT m.order_id, m.order_total, oi.sku
                FROM ml_devolucoes m
                LEFT JOIN order_items oi ON oi.order_id = m.order_id::text
                WHERE m.order_id = ANY(%s)
                  AND m.claim_type = 'mediations' AND m.claim_status = 'closed'
            """, conn, params=(todo,))
            if not df_db.empty:
                db_map = {int(r["order_id"]): r.to_dict() for _, r in df_db.iterrows()}
        except Exception:
            pass
        try:
            df_tiny = pd.read_sql(
                "SELECT sku, preco_custo, nome FROM tiny_sku_costs WHERE preco_custo > 0", conn
            )
            tiny_map = {
                str(r["sku"]).upper(): (float(r["preco_custo"]), str(r["nome"]))
                for _, r in df_tiny.iterrows()
            } if not df_tiny.empty else {}
        except Exception:
            pass

    # 4. Função de validação de um pedido (só I/O externo, sem DB)
    def validate_one(order_id: int) -> dict:
        mp_info  = mp_rows.get(order_id, {})
        db_info  = db_map.get(order_id, {})
        db_total = float(db_info.get("order_total") or mp_info.get("mp_valor") or 0)
        sku_raw  = str(db_info.get("sku") or "").strip()
        sku_up   = sku_raw.upper()

        tiny_cmv = tiny_nome = None
        if sku_up and sku_up in tiny_map:
            tiny_cmv, tiny_nome = tiny_map[sku_up]
        elif sku_up:
            try:
                from src.api import tiny_client as _tc
                prod = _tc.lookup_by_sku(sku_raw)
                if prod:
                    tiny_cmv  = float(prod.get("preco_custo") or 0) or None
                    tiny_nome = str(prod.get("nome") or "") or None
            except Exception:
                pass

        base = {
            "order_id":        order_id,
            "db_total":        db_total,
            "api_total":       None,
            "delta":           None,
            "api_pay_status":  None,
            "api_pay_detail":  None,
            "concorda_ml":     False,
            "mp_status_detail": mp_info.get("status_detail"),
            "mp_valor":        mp_info.get("mp_valor"),
            "sku":             sku_raw or None,
            "tiny_cmv":        tiny_cmv,
            "tiny_nome":       tiny_nome,
        }
        try:
            order = ml_client.get_order(order_id)
            if not order:
                return base
            api_total  = float(order.get("total_amount") or 0)
            pays       = order.get("payments") or [{}]
            pay0       = pays[0] if pays else {}
            base["api_total"]      = api_total
            base["delta"]         = abs(api_total - db_total)
            base["api_pay_status"]= pay0.get("status", "")
            base["api_pay_detail"]= pay0.get("status_detail", "")
            base["concorda_ml"]   = base["delta"] < 1.0
        except Exception as exc:
            base["api_pay_detail"] = str(exc)[:120]
        return base

    # 5. Validar em paralelo, coletar resultados
    results: list[dict] = []
    lock   = threading.Lock()
    done   = [0]
    n_todo = len(todo)

    if todo:
        with ThreadPoolExecutor(max_workers=n_workers) as ex:
            futs = {ex.submit(validate_one, oid): oid for oid in todo}
            for fut in as_completed(futs):
                r = fut.result()
                with lock:
                    results.append(r)
                    done[0] += 1
                    if done[0] % 200 == 0 or done[0] == n_todo:
                        print(f"    {done[0]:,}/{n_todo:,} validados...", flush=True)

        # 6. Bulk UPSERT — usa conexão fresca (a original pode ter expirado em runs longos)
        if results:
            insert_conn = get_db_connection()
            try:
                _ensure_validation_table(insert_conn)
                CHUNK = 200
                for start in range(0, len(results), CHUNK):
                    batch = results[start:start + CHUNK]
                    with insert_conn.cursor() as cur:
                        for r in batch:
                            cur.execute("""
                                INSERT INTO mp_validation_results
                                    (order_id, db_total, api_total, delta, api_pay_status,
                                     api_pay_detail, concorda_ml, mp_status_detail, mp_valor,
                                     sku, tiny_cmv, tiny_nome)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                                ON CONFLICT (order_id) DO UPDATE SET
                                    validated_at    = NOW(),
                                    db_total        = EXCLUDED.db_total,
                                    api_total       = EXCLUDED.api_total,
                                    delta           = EXCLUDED.delta,
                                    api_pay_status  = EXCLUDED.api_pay_status,
                                    api_pay_detail  = EXCLUDED.api_pay_detail,
                                    concorda_ml     = EXCLUDED.concorda_ml,
                                    mp_status_detail= EXCLUDED.mp_status_detail,
                                    mp_valor        = EXCLUDED.mp_valor,
                                    sku             = EXCLUDED.sku,
                                    tiny_cmv        = EXCLUDED.tiny_cmv,
                                    tiny_nome       = EXCLUDED.tiny_nome
                            """, (
                                r["order_id"], r["db_total"], r["api_total"], r["delta"],
                                r["api_pay_status"], r["api_pay_detail"], r["concorda_ml"],
                                r["mp_status_detail"], r["mp_valor"], r["sku"],
                                r["tiny_cmv"], r["tiny_nome"],
                            ))
                    insert_conn.commit()
                    print(f"    Salvos {min(start+CHUNK, len(results)):,}/{len(results):,} no Neon...", flush=True)
            finally:
                insert_conn.close()

    # 7. Ler TODOS do cache (usando conexão fresca para garantir resultados atualizados)
    read_conn = get_db_connection()
    try:
        df_all = pd.read_sql("""
            SELECT order_id, db_total, api_total, delta, api_pay_status, api_pay_detail,
                   concorda_ml, mp_status_detail, mp_valor, sku, tiny_cmv, tiny_nome
            FROM mp_validation_results
            ORDER BY delta DESC NULLS LAST
        """, read_conn)
    finally:
        read_conn.close()

    n_total    = len(df_all)
    n_concorda = int(df_all["concorda_ml"].sum()) if n_total > 0 else 0
    n_diverge  = n_total - n_concorda
    n_cmv      = int(df_all["tiny_cmv"].notna().sum()) if n_total > 0 else 0
    pct        = n_concorda / n_total * 100 if n_total > 0 else 0.0

    return {
        "verificados":  n_total,
        "concordantes": n_concorda,
        "divergem":     n_diverge,
        "com_cmv":      n_cmv,
        "pct_paridade": round(pct, 1),
        "novos":        len(results),
        "detalhes":     df_all.to_dict("records") if n_total > 0 else [],
    }


# ─── Função principal ─────────────────────────────────────────────────────────

def main(pasta: Path, output_dir: Path, force: bool = False, revalidar: bool = False, validar_base: bool = False) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    print(SEP)
    print("PROCESSADOR DE RELATÓRIOS MP – Análise Financeira de Devoluções")
    print(f"Iniciado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print(SEP)

    # ── 1. Ingestão ───────────────────────────────────────────────────────────
    print(f"\n[1/4] Ingerindo relatórios MP de: {pasta}")
    resultados = scan_folder(pasta, force=force)
    if resultados:
        for r in resultados:
            if r.get("ja_importado"):
                print(f"  - {r['filename']} → já importado (use --force para reimportar)")
            elif "erro" in r:
                print(f"  ✗ {r['filename']} → ERRO: {r['erro']}")
            else:
                print(f"  ✓ {r['filename']} → {r['rows_new']} novos | {r['rows_dup']} duplicados")
    else:
        print(f"  ⚠ Nenhum arquivo after_collection encontrado em {pasta}")

    # ── 2. Carga do Neon + Conciliação completa ──────────────────────────────
    print("\n[2/5] Carregando dados do Neon + conciliação completa ML API + Tiny...")
    conn = get_db_connection()
    try:
        # valida primeiro: a carga usa mp_validation_results fresco na conciliação
        validation = _run_full_validation(conn, force=revalidar, n_workers=8, validar_base_completa=validar_base)
        df_mp, df_port = _load_and_enrich(conn)
    finally:
        conn.close()
    print(f"  ✓ Concordância ML API: {validation['pct_paridade']}%  ({validation['concordantes']:,}/{validation['verificados']:,} pedidos | {validation['novos']:,} novos | {validation['com_cmv']:,} c/ CMV Tiny)")

    if df_mp.empty:
        print("  ⚠ Nenhuma transação MP encontrada no Neon. Ingira pelo menos um arquivo primeiro.")
        return

    print(f"  ✓ {df_mp['mp_order_id'].nunique():,} pedidos no relatório MP")
    print(f"  ✓ {df_port['order_id'].nunique():,} pedidos no portfólio Neon")
    print(f"  ✓ {df_port[df_port['perda_bruta']>0]['order_id'].nunique():,} pedidos com devolução/mediação")

    kpi = _kpis(df_mp, df_port)

    # ── 3. Relatórios ─────────────────────────────────────────────────────────
    print("\n[3/5] Gerando relatórios...")
    html_out = output_dir / f"relatorio_devolucoes_{TODAY}.html"
    xlsx_out = output_dir / f"relatorio_devolucoes_{TODAY}.xlsx"

    _build_html_dashboard(df_mp, df_port, kpi, html_out, validation=validation)
    _build_xlsx(df_mp, df_port, kpi, xlsx_out)

    # JSON de metadados
    json_out = output_dir / f"relatorio_devolucoes_{TODAY}.json"
    json_out.write_text(json.dumps({
        "gerado_em": datetime.now().isoformat(),
        "kpi": {k: (round(float(v), 2) if isinstance(v, (int, float)) and not isinstance(v, bool) else v)
                for k, v in kpi.items()},
        "validacao": {k: v for k, v in validation.items() if k != "detalhes"},
        "ingestao": resultados,
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  ✓ JSON metadata salvo em: {json_out}")

    # Cópia com nome estável — é este arquivo que o ml_live_poll.py mantém
    # atualizado e que o navegador acompanha (banner de atualização automática).
    import shutil
    live_html = output_dir / "painel_devolucoes_live.html"
    live_json = output_dir / "painel_devolucoes_live.json"
    shutil.copyfile(html_out, live_html)
    shutil.copyfile(json_out, live_json)
    print(f"  ✓ Painel vivo: {live_html}")

    # ── 4. Resumo ─────────────────────────────────────────────────────────────
    print(f"\n[4/5] Resumo Financeiro")
    print(SEP)
    print(f"  Portfólio : {kpi['total_pedidos']:,} pedidos")
    print(f"  Taxa ML   : {BRL(kpi['taxa_ml'])}")
    print(f"  Tarifa frete: {BRL(kpi['tarifa_frete'])}  |  Frete cobrado: {BRL(kpi['frete_cobrado'])}  |  Saldo: {BRL(kpi['saldo_frete'])}")
    print(f"  Cancelamentos: {BRL(kpi['cancelamentos'])}")
    print()
    print(f"  Devoluções: {kpi['dev_pedidos']:,} pedidos ({PCT(kpi['dev_pct_portfolio'])} do portfólio)")
    print(f"  Perda bruta  : {BRL(kpi['perda_bruta'])}")
    print(f"  Recuperado ML: {BRL(kpi['recuperado_ml'])}  ({PCT(kpi['pct_recuperado'])})")
    print(f"  Taxa retida  : {BRL(kpi['taxa_retida'])}")
    print(f"  PERDA LÍQUIDA: {BRL(kpi['perda_liquida'])}")
    print()
    print(f"  MP Report  : {kpi['mp_total_tx']:,} transações | {kpi['mp_pedidos_unicos']:,} pedidos | {BRL(kpi['mp_valor_total'])}")
    print()
    print(f"  Concordância ML API: {validation['pct_paridade']}%  ({validation['concordantes']:,}/{validation['verificados']:,} pedidos) | CMV Tiny: {validation['com_cmv']:,} SKUs")
    print(SEP)
    print(f"\n  Arquivos gerados:")
    print(f"    HTML : {html_out}")
    print(f"    XLSX : {xlsx_out}")
    print(f"    JSON : {json_out}")


# ─── CLI ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    ap = argparse.ArgumentParser(
        description="Processa relatórios MP e gera análise financeira de devoluções."
    )
    ap.add_argument(
        "--pasta",
        default=str(DEFAULT_IN),
        help=f"Pasta com os arquivos after_collection*.xlsx (padrão: {DEFAULT_IN})",
    )
    ap.add_argument(
        "--output",
        default=str(DEFAULT_OUT),
        help=f"Pasta de saída dos relatórios (padrão: {DEFAULT_OUT})",
    )
    ap.add_argument(
        "--force",
        action="store_true",
        help="Reimportar arquivos já existentes no log",
    )
    ap.add_argument(
        "--revalidar",
        action="store_true",
        help="Forçar re-validação de todos os pedidos (ignora cache Neon)",
    )
    ap.add_argument(
        "--validar-base-completa",
        action="store_true",
        dest="validar_base",
        help="Incluir toda a base ml_devolucoes (13+ meses, ~10 min na 1ª execução) na conciliação",
    )
    args = ap.parse_args()
    main(Path(args.pasta), Path(args.output), force=args.force, revalidar=args.revalidar, validar_base=args.validar_base)
